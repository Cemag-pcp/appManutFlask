# app.py
from flask import Flask, current_app, send_file, jsonify, render_template, request, redirect, url_for, flash, Blueprint, Response, send_from_directory, send_file, current_app, make_response
import psycopg2  # pip install psycopg2
import psycopg2.extras
import datetime
import pandas as pd
import numpy as np
import json
from funcoes import gerador_de_semanas_informar_manutencao, login_required, gerador_de_semanas_informar_manutencao_diario,gerar_planejamento_maquinas_preventivas,calcular_proxima_data
import warnings
from flask import session
import base64
from datetime import datetime, timedelta, time, date
from pandas.tseries.offsets import BMonthEnd,MonthEnd,BDay
from psycopg2 import Error
import json
from PIL import Image
import io
from openpyxl import load_workbook
# import convertapi
from werkzeug.utils import secure_filename
import os
import zipfile
from io import BytesIO
import re

routes_bp = Blueprint('routes', __name__)

# routes_bp.config['UPLOAD_FOLDER'] = r'C:\Users\pcp2\projetoManutencao\appManutFlask-3\UPLOAD_FOLDER'

# Configurar a pasta para salvar os vídeos
routes_bp.config = {}
routes_bp.config['UPLOAD_FOLDER'] = 'UPLOAD_FOLDER'

warnings.filterwarnings("ignore")

# DB_HOST = "localhost"
DB_HOST = "database-1.cdcogkfzajf0.us-east-1.rds.amazonaws.com"
DB_NAME = "postgres"
DB_USER = "postgres"
DB_PASS = "15512332"

conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)

def dados_para_editar(id_ordem,n_ordem):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = "select o.*,func.nome,func.matricula,coalesce(status,'Em espera') as status_new from tb_ordens as o left join tb_funcionario as func on ',' || o.operador || ',' like '%%,' || func.matricula || ',%%' where id_ordem = %s and n_ordem = %s"
    sql_data_abertura =  """select dataabertura - INTERVAL '3 hours' as dataabertura,solicitante,
                            equipamento_em_falha,cod_equipamento,setor_maquina_solda,status
                            from tb_ordens where id_ordem = %s and n_ordem = 0
                        """
    sql_tombamento = """select tombamento from tb_ordens
                        left join tb_maquinas as t2 on t2.codigo = maquina
                        where id_ordem = %s
                        limit 1
                    """
    # sql_maquina_preventiva = """select codigo from tb_maquinas_preventivas where codigo = %s"""
    
    # Informações gerais
    cur.execute(sql,(id_ordem,n_ordem))
    data = cur.fetchall()
    
    # dataabertura
    cur.execute(sql_data_abertura,(id_ordem,))
    dados_dataabertura = cur.fetchall()

    # Tombamento
    cur.execute(sql_tombamento,(id_ordem,))
    tombamento = cur.fetchall()

    setor_values = [row['setor'] for row in data][0]
    solicitante = dados_dataabertura[0]['solicitante']
    data_abertura = dados_dataabertura[0]['dataabertura']
    n_execucao = [row['n_ordem'] for row in data][0]
    maquina = [row['maquina'] for row in data][0]
    tombamento = [row['tombamento'] for row in tombamento][0]
    equipamento_em_falha = dados_dataabertura[0]['equipamento_em_falha']
    codigo_equipamento = dados_dataabertura[0]['cod_equipamento']
    setor_maquina_solda = dados_dataabertura[0]['setor_maquina_solda']
    risco = [row['risco'] for row in data][0]
    desc_usuario = [row['problemaaparente'] for row in data][0]
    status = [row['status_new'] for row in data][0]
    tipo_manutencao = [row['tipo_manutencao'] for row in data][0]
    area_manutencao = [row['area_manutencao'] for row in data][0]
    pvlye = [row['pvlye'] for row in data][0]
    paplus = [row['pa_plus'] for row in data][0]
    tratamento = [row['tratamento'] for row in data][0]
    phagua = [row['ph_agua'] for row in data][0]

    try:
        operador = [row['matricula'] + " - " + row['nome'] for row in data]
    except TypeError:
        operador = ''

    descmanutencao = [row['descmanutencao'] for row in data][0]
    # data_inicio_fim = [row['datainicio'] for row in data][0]
    datainicio = [row['datainicio'] for row in data][0]
    horainicio = [row['horainicio'] for row in data][0]
    datafim = [row['datafim'] for row in data][0]
    horafim = [row['horafim'] for row in data][0]

    # Formatando as datas e horas como strings
    data_inicio_str = datainicio.isoformat()
    hora_inicio_str = horainicio.strftime('%H:%M')
    data_fim_str = datafim.isoformat()
    hora_fim_str = horafim.strftime('%H:%M')

    dados = {
        'setor_values':setor_values,
        'solicitante':solicitante,
        'data_abertura':data_abertura,
        'n_execucao':n_execucao,
        'maquina':maquina,
        'tombamento':tombamento,
        'equipamento_em_falha':equipamento_em_falha,
        'codigo_equipamento':codigo_equipamento,
        'setor_maquina_solda':setor_maquina_solda,
        'risco':risco,
        'desc_usuario':desc_usuario,
        'status':status,
        'tipo_manutencao':tipo_manutencao,
        'area_manutencao':area_manutencao,
        'operador':operador,
        'descmanutencao':descmanutencao,
        'pvlye':pvlye,
        'paplus':paplus,
        'tratamento':tratamento,
        'phagua':phagua,
        'data_inicio': data_inicio_str,
        'hora_inicio': hora_inicio_str,
        'data_fim': data_fim_str,
        'hora_fim': hora_fim_str
    }

    return dados

def buscar_dados_os(id_ordem):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = "select *, coalesce(status,'Em espera') as status_new from tb_ordens where id_ordem = %s order by n_ordem desc limit 1"
    sql_data_abertura =  """select dataabertura - INTERVAL '3 hours' as dataabertura,solicitante,
                            equipamento_em_falha,cod_equipamento,setor_maquina_solda,status
                            from tb_ordens where id_ordem = %s and n_ordem = 0
                        """
    sql_tombamento = """select tombamento from tb_ordens
                        left join tb_maquinas as t2 on t2.codigo = maquina
                        where id_ordem = %s
                        limit 1
                    """
    sql_maquina_preventiva = """select codigo from tb_maquinas_preventivas where codigo = %s"""
    
    # Informações gerais
    cur.execute(sql,(id_ordem,))
    data = cur.fetchall()
    
    # dataabertura
    cur.execute(sql_data_abertura,(id_ordem,))
    dados_dataabertura = cur.fetchall()

    # Tombamento
    cur.execute(sql_tombamento,(id_ordem,))
    tombamento = cur.fetchall()

    setor_values = [row['setor'] for row in data][0]
    solicitante = dados_dataabertura[0]['solicitante']
    data_abertura = dados_dataabertura[0]['dataabertura']
    n_execucao = [row['n_ordem'] for row in data][0]+1
    maquina = [row['maquina'] for row in data][0]
    tombamento = [row['tombamento'] for row in tombamento][0]
    equipamento_em_falha = dados_dataabertura[0]['equipamento_em_falha']
    codigo_equipamento = dados_dataabertura[0]['cod_equipamento']
    setor_maquina_solda = dados_dataabertura[0]['setor_maquina_solda']
    risco = [row['risco'] for row in data][0]
    desc_usuario = [row['problemaaparente'] for row in data][0]
    status = [row['status_new'] for row in data][0]
    tipo_manutencao = [row['tipo_manutencao'] for row in data][0]
    area_manutencao = [row['area_manutencao'] for row in data][0]

    # Maquina preventiva
    cur.execute(sql_maquina_preventiva,(maquina,))
    preventiva = cur.fetchall()
    
    if len(preventiva) > 0:
        preventiva = True
    else:
        preventiva = False

    lista_opcoes = ['Em execução', 'Finalizada', 'Aguardando material']

    opcoes = []
    opcoes.append(status)

    for opcao in lista_opcoes:
        opcoes.append(opcao)

    opcoes = list(set(opcoes))
    opcoes.remove(status)  # Remove o elemento 'c' da lista
    opcoes.insert(0, status)

    dados = {
        'setor_values':setor_values,
        'solicitante':solicitante,
        'data_abertura':data_abertura,
        'n_execucao':n_execucao,
        'maquina':maquina,
        'tombamento':tombamento,
        'equipamento_em_falha':equipamento_em_falha,
        'codigo_equipamento':codigo_equipamento,
        'setor_maquina_solda':setor_maquina_solda,
        'risco':risco,
        'desc_usuario':desc_usuario,
        'status':status,
        'tipo_manutencao':tipo_manutencao,
        'area_manutencao':area_manutencao,
        'preventiva':preventiva,
        'opcoes':opcoes
    }

    return dados

def buscar_funcionarios():
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """SELECT * FROM tb_funcionario"""
    tb_funcionarios = pd.read_sql_query(query, conn)
    tb_funcionarios['matricula_nome'] = tb_funcionarios['matricula'] + \
        " - " + tb_funcionarios['nome']
    funcionarios = tb_funcionarios[['matricula_nome']].values.tolist()

    return funcionarios


def calcular_minutos_uteis(row, df):

    """
    Função para calcular os minutos úteis entre duas datas
    """

    hora_inicio_trabalho = 7
    hora_fim_trabalho = 17

    # Extraia as datas de início e fim da linha
    if row['parada3'] == 'true':
        data_inicio = pd.to_datetime(row['inicio']).replace(tzinfo=None)  # Remova as informações de fuso horário
        data_fim = row['fim'].replace(tzinfo=None)  # Remova as informações de fuso horário
    else:
        return 0
    
    def dentro_do_horario_trabalho(data):
        return hora_inicio_trabalho <= data.hour < hora_fim_trabalho

    # Inicialize variáveis
    minutos_uteis = 0
    data_atual = data_inicio

    # Calcule o número de minutos úteis
    while data_atual <= data_fim:
        if data_atual.weekday() < 5 and dentro_do_horario_trabalho(data_atual):
            minutos_uteis += 1
        data_atual += timedelta(minutes=1)

    return minutos_uteis


def obter_nome_mes(numeros_meses):

    """
    Função para obter nome do mês com base no número dele.
    """

    numeros_meses = list(map(int, numeros_meses))

    nomes_meses = {
        1: 'Janeiro',
        2: 'Fevereiro',
        3: 'Março',
        4: 'Abril',
        5: 'Maio',
        6: 'Junho',
        7: 'Julho',
        8: 'Agosto',
        9: 'Setembro',
        10: 'Outubro',
        11: 'Novembro',
        12: 'Dezembro'
    }

    nomes = [nomes_meses.get(numero_mes, '') for numero_mes in numeros_meses]

    return nomes


def ultimo_dia_mes(mes):

    """
    Função para buscar o último dia do mês e hora.
    """

    # Especifique o mês desejado
    mes_desejado = mes[-1]  # Outubro

    print(mes_desejado)
    # Obtenha o último dia do mês desejado
    if mes_desejado == 12:
        ultimo_dia_do_mes = datetime(datetime.now().year, mes_desejado, 1) - timedelta(days=1)
    elif mes_desejado != 1:
        ultimo_dia_do_mes = datetime(2023, mes_desejado + 1, 1) - timedelta(days=1)
    else:
        ultimo_dia_do_mes = datetime(datetime.now().year, mes_desejado + 1, 1) - timedelta(days=1)

    print(datetime.now().year)

    # Defina a hora desejada
    hora_desejada = time(23, 59, 59)

    # Combine a data e a hora
    ultimo_dia_do_mes_com_hora = datetime.combine(ultimo_dia_do_mes, hora_desejada)

    return ultimo_dia_do_mes_com_hora


def dias_uteis(meses):

    """
    Função que calcula os dias úteis em um determinado período de meses.
    """

    qtd_dias_uteis_total = 0

    for mes in meses:
        # Verificar se o mês é válido (entre 1 e 12)
        if not mes:
            data_atual = pd.Timestamp.now()
            mes = data_atual.month  # Mês atual
            qtd_dias_uteis = 0

            for m in range(7, mes + 1):
                primeiro_dia_mes = pd.Timestamp(data_atual.year, m, 1)

                if m == mes:
                    # Se for o mês atual, use o dia atual como o último dia útil
                    ultimo_dia_util_mes = data_atual
                else:
                    # Se não for o mês atual, use o último dia útil do mês
                    ultimo_dia_util_mes = primeiro_dia_mes + pd.offsets.BMonthEnd()

                datas_uteis = pd.bdate_range(
                    primeiro_dia_mes, ultimo_dia_util_mes)
                qtd_dias_uteis += len(datas_uteis)

        elif mes == int(pd.Timestamp.now().month):
            # Obter a data atual
            data_atual = pd.Timestamp.now()

            # Obter o primeiro dia do mês atual
            primeiro_dia_mes = data_atual - pd.offsets.MonthBegin()

            # Obter o último dia útil do mês atual
            ultimo_dia_util_mes = primeiro_dia_mes + BMonthEnd()

            # Obter a sequência de datas úteis no mês atual
            datas_uteis = pd.bdate_range(primeiro_dia_mes, data_atual)

            # Contar o número de dias úteis
            qtd_dias_uteis = len(datas_uteis)

        else:
            data_atual = pd.Timestamp.now()
            primeiro_dia_mes = pd.Timestamp(data_atual.year, mes, 1)

            # Obter o último dia útil do mês especificado
            ultimo_dia_util_mes = primeiro_dia_mes + BMonthEnd()

            # Obter a sequência de datas úteis no mês especificado
            datas_uteis = pd.bdate_range(primeiro_dia_mes, ultimo_dia_util_mes)

            # Contar o número de dias úteis
            qtd_dias_uteis = len(datas_uteis)

        qtd_dias_uteis_total += qtd_dias_uteis

    return qtd_dias_uteis_total

def dias_uteis_inicial_final(dia_inicial,dia_final):

    # Convertendo as strings para objetos de data
    data_inicial = pd.to_datetime(dia_inicial)
    data_final = pd.to_datetime(dia_final)

    # Criando um intervalo de datas entre as datas inicial e final
    intervalo = pd.date_range(start=data_inicial, end=data_final)

    # Contando apenas os dias úteis no intervalo
    dias_uteis = len([dia for dia in intervalo if dia.weekday() < 5])

    return dias_uteis

def tempo_os():

    """
    Função para calcular tempo de execução de os. 
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    # Obtém os dados da tabela
    s = ("""
        SELECT datafim,
            TO_TIMESTAMP(datainicio || ' ' || horainicio, 'YYYY-MM-DD HH24:MI:SS') AS inicio,
            TO_TIMESTAMP(datafim || ' ' || horafim, 'YYYY-MM-DD HH24:MI:SS') AS fim
        FROM tb_ordens
    """)

    df_timeline = pd.read_sql_query(s, conn)

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    df_timeline = df_timeline.dropna()

    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
            lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    df_timeline = df_timeline[['datafim', 'diferenca']]
    df_agrupado = df_timeline.groupby(
        'datafim')['diferenca'].sum().reset_index()

    # df_timeline = df_timeline.values.tolist()

    return df_agrupado


def cards_get(query):

    """
    Função para gerar dados de quantidade de os em aberto, em execução, aguardadno material e fechada.
    """
    
    query = """
        SELECT *
        FROM tb_ordens
        WHERE (ordem_excluida IS NULL OR ordem_excluida = FALSE)
        """
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cards = pd.read_sql_query(query, conn)
    # cards = cards[cards['id_ordem'] == 837]

    cards['status'] = cards['status'].fillna('Em espera')

    cards = cards.sort_values(by='n_ordem', ascending=True)

    cards = cards.drop_duplicates(subset='id_ordem', keep='last')

    em_execucao = cards[cards['status'] == 'Em espera'][['id_ordem', 'status','n_ordem']]

    print(em_execucao)

    cards = cards.groupby(['status'])['status'].count()

    # Crie um dicionário para armazenar os resultados
    status_dict = {}
    for status, qt_os in cards.items():
        status_dict[status] = qt_os

    # Certifique-se de que todas as chaves estão presentes no dicionário, mesmo que com valor 0
    lista_qt = [
        status_dict.get('Aguardando material', 0),
        status_dict.get('Finalizada', 0),
        status_dict.get('Em execução', 0),
        status_dict.get('Aguardando OK', 0),
        status_dict.get('Em espera', 0)
    ]

    print(lista_qt)

    return lista_qt


def cards_post(query):

    """
    Função para gerar dados de quantidade de os em aberto, em execução, aguardadno material e fechada.
    """
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cards = pd.read_sql_query(query, conn)
    # cards = cards[cards['id_ordem'] == 837]

    cards['status'] = cards['status'].fillna('Em espera')

    cards = cards.sort_values(by='n_ordem', ascending=True)

    cards = cards.drop_duplicates(subset='id_ordem', keep='last')

    em_execucao = cards[cards['status'] == 'Em espera'][['id_ordem', 'status','n_ordem']]

    print(em_execucao)

    cards = cards.groupby(['status'])['status'].count()

    # Crie um dicionário para armazenar os resultados
    status_dict = {}
    for status, qt_os in cards.items():
        status_dict[status] = qt_os

    # Certifique-se de que todas as chaves estão presentes no dicionário, mesmo que com valor 0
    lista_qt = [
        # status_dict.get('Em espera', 0),
        status_dict.get('Aguardando material', 0),
        status_dict.get('Finalizada', 0),
        status_dict.get('Em execução', 0),
        status_dict.get('Aguardando OK', 0)
    ]

    print(lista_qt)

    return lista_qt

def card_post_em_espera(query_em_espera):

    """
    Função para gerar dados de quantidade de os em aberto, em execução, aguardadno material e fechada.
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cards = pd.read_sql_query(query_em_espera, conn)
    # cards = cards[cards['id_ordem'] == 837]

    cards['status'] = cards['status'].fillna('Em espera')

    cards = cards.sort_values(by='n_ordem', ascending=True)

    print(cards)

    cards = cards.drop_duplicates(subset='id_ordem', keep='last')

    print(cards)

    em_execucao = cards[cards['status'] == 'Em espera'][['id_ordem', 'status','n_ordem']]

    print(em_execucao)

    cards = cards.groupby(['status'])['status'].count()

    # Crie um dicionário para armazenar os resultados
    status_dict = {}
    for status, qt_os in cards.items():
        status_dict[status] = qt_os

    # Certifique-se de que todas as chaves estão presentes no dicionário, mesmo que com valor 0
    card_em_espera = [
        status_dict.get('Em espera', 0),
    ]

    print(card_em_espera)

    return card_em_espera

def funcao_geral(query_mtbf, query_mttr, boleano_historico, setor_selecionado, query_disponibilidade, query_horas_trabalhada_tipo, query_horas_trabalhada_area, query_horas_trabalhada_setor, dia_inicial,dia_final,lista_meses):

    """
    Função para gerar todos os gráficos
    """
    if setor_selecionado == '':
        setor_selecionado = None

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    """
    Gráfico de MTBF por máquina
    utiliza a query_mtbf
    """

    # query_mtbf
    # mtbf_maquina
    # Obtém os dados da tabela

    df_timeline = pd.read_sql_query(query_mtbf, conn)

    df_timeline['datafim'] = pd.to_datetime(df_timeline['datafim'])
    # Crie uma nova coluna 'mes' usando o atributo .dt.month
    df_timeline['mes'] = df_timeline['datafim'].dt.month

    # mes_hoje = datetime.today().month - 1
    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    df_timeline = df_timeline.dropna()

    df_timeline['maquina'] = df_timeline['maquina']
    df_timeline['maquina'] = df_timeline['maquina'].str.split(' - ').str[0]

    # Contar a quantidade de manutenções por máquina
    contagem = df_timeline['maquina'].value_counts()
    df_timeline['qtd_manutencao'] = df_timeline['maquina'].map(contagem)
    df_timeline = df_timeline.drop_duplicates(subset='maquina')

    qtd_dias_uteis = dias_uteis_inicial_final(dia_inicial,dia_final)

    df_timeline['carga_trabalhada'] = qtd_dias_uteis * 9

    df_timeline['MTBF'] = ((df_timeline['carga_trabalhada']) /
                           df_timeline['qtd_manutencao']).round(2)

    df_timeline.sort_values(by='MTBF', inplace=True)

    df_timeline = df_timeline[['maquina',
                               'qtd_manutencao', 'carga_trabalhada', 'MTBF']]

    def convert_time_to_decimal(time_str):
        hours, minutes, seconds = map(int, time_str.split(":"))
        return hours + minutes / 60 + seconds / 3600

    if len(df_timeline) > 0:

        if boleano_historico == True:

            """
            Se for GET pega todo o histórico e adiciona no atual
            """
            
            historico_csv = pd.read_csv("mtbf_historico.csv", sep=';')
            historico_csv["historico_mtbf_decimal"] = historico_csv["historico_mtbf"].apply(
                convert_time_to_decimal)

            # if setor_selecionado:
            #     historico_csv = historico_csv[historico_csv['setor']
            #                                   == setor_selecionado]

            historico_csv['maquina'] = historico_csv['maquina'].str.split(
                ' - ').str[0]

            df_timeline = df_timeline.merge(
                historico_csv, how='outer', on='maquina').fillna(0)
            df_timeline['MTBF_final'] = df_timeline['historico_mtbf_decimal'] + \
                df_timeline['MTBF']
            df_timeline = df_timeline.drop(columns={'MTBF'})
            df_timeline = df_timeline.rename(
                columns={"MTBF_final": 'MTBF'}).round(2)
            df_timeline.sort_values("MTBF", inplace=True)
            media_qtd_manutencao = df_timeline['qtd_manutencao'][df_timeline['qtd_manutencao'] != 0].mean()

            # Substitui os valores 0 em qtd_manutencao pela média calculada
            df_timeline['qtd_manutencao'] = df_timeline['qtd_manutencao'].replace(0,round(media_qtd_manutencao))

            df_timeline['carga_trabalhada'] = qtd_dias_uteis * 9

            grafico1_maquina = df_timeline['maquina'].tolist()  # eixo x
            grafico1_mtbf = df_timeline['MTBF'].tolist()  # eixo y gráfico 1

            df_timeline_copia = df_timeline[[
                'maquina', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        else:

            grafico1_maquina = df_timeline['maquina'].tolist()  # eixo x
            grafico1_mtbf = df_timeline['MTBF'].tolist()  # eixo y gráfico 1

            df_timeline_copia = df_timeline[[
                'maquina', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        context_mtbf_maquina = {
            'labels_mtbf_maquina': grafico1_maquina, 'dados_mtbf_maquina': grafico1_mtbf}

    else:

        grafico1_maquina = []
        grafico1_mtbf = []
        df_timeline_copia = df_timeline[[
            'maquina', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        context_mtbf_maquina = {
            'labels_mtbf_maquina': grafico1_maquina, 'dados_mtbf_maquina': grafico1_mtbf}

    """
    Gráfico de MTBF por setor
    utiliza a query_mtbf
    """

    # query_mtbf
    # mtbf_setor

    df_timeline = pd.read_sql_query(query_mtbf, conn)

    # df_timeline = df_timeline[df_timeline['n_ordem'] == 0]

    df_timeline['datafim'] = pd.to_datetime(df_timeline['datafim'])
    # Crie uma nova coluna 'mes' usando o atributo .dt.month
    df_timeline['mes'] = df_timeline['datafim'].dt.month

    # mes_hoje = datetime.today().month - 1
    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    df_timeline = df_timeline.dropna()

    df_timeline['setor'].value_counts()

    # Contar a quantidade de manutenções por máquina
    contagem = df_timeline['setor'].value_counts()
    df_timeline['qtd_manutencao'] = df_timeline['setor'].map(contagem)
    df_timeline = df_timeline.drop_duplicates(subset='setor')

    qtd_dias_uteis = dias_uteis_inicial_final(dia_inicial,dia_final)

    df_timeline['carga_trabalhada'] = qtd_dias_uteis * 9

    df_timeline['MTBF'] = ((df_timeline['carga_trabalhada']) /
                           df_timeline['qtd_manutencao']).round(2)

    df_timeline.sort_values(by='MTBF', inplace=True)

    df_timeline_mtbf_setor = df_timeline[[
        'setor', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']]

    if len(df_timeline) > 0:

        if boleano_historico == True:
            """
            Se for GET pega todo o histórico e adiciona no atual
            """

            historico_csv = pd.read_csv("mtbf_historico.csv", sep=';')
            historico_csv["historico_mtbf_decimal"] = historico_csv["historico_mtbf"].apply(
                convert_time_to_decimal)
            historico_csv = historico_csv.groupby(
                ['setor']).sum().reset_index()

            df_timeline_mtbf_setor = df_timeline_mtbf_setor.merge(
                historico_csv, how='outer', on='setor').fillna(0)
            df_timeline_mtbf_setor['MTBF_final'] = df_timeline_mtbf_setor['historico_mtbf_decimal'] + \
                df_timeline_mtbf_setor['MTBF']
            df_timeline_mtbf_setor = df_timeline_mtbf_setor.drop(columns={
                                                                 'MTBF'})
            df_timeline_mtbf_setor = df_timeline_mtbf_setor.rename(
                columns={"MTBF_final": 'MTBF'}).round(2)
            df_timeline_mtbf_setor.sort_values("MTBF", inplace=True)

            # eixo x
            grafico1_maquina = df_timeline_mtbf_setor['setor'].tolist()
            # eixo y gráfico 1
            grafico1_mtbf = df_timeline_mtbf_setor['MTBF'].tolist()

            df_timeline_mtbf_setor = df_timeline_mtbf_setor[[
                'setor', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        else:

            grafico1_maquina = df_timeline['setor'].tolist()  # eixo x
            grafico1_mtbf = df_timeline['MTBF'].tolist()  # eixo y gráfico 1

            df_timeline_mtbf_setor = df_timeline[[
                'setor', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        context_mtbf_setor = {
            'labels_mtbf_setor': grafico1_maquina, 'dados_mtbf_setor': grafico1_mtbf}

    else:

        grafico1_maquina = []
        grafico1_mtbf = []
        df_timeline_mtbf_setor = df_timeline[[
            'setor', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        context_mtbf_setor = {
            'labels_mtbf_setor': grafico1_maquina, 'dados_mtbf_setor': grafico1_mtbf}

    """
    Gráfico de MTTR por máquina
    utiliza a query_mttr
    """

    # query_mttr
    # mttr_maquina

    df_timeline = pd.read_sql_query(query_mttr, conn)

    df_timeline = df_timeline.sort_values(by='n_ordem')

    df_timeline = df_timeline.dropna()

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    df_timeline = df_timeline.dropna()
    
    df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])
    # Crie uma nova coluna 'mes' usando o atributo .dt.month
    df_timeline['mes'] = df_timeline['fim'].dt.month

    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    df_timeline['now'] = datetime.now()
    df_timeline['mes_hoje'] = datetime.now().month
    # df_timeline['ultimo_dia_mes'] = ultimo_dia_mes(lista_meses)

    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = df_timeline.apply(calcular_minutos_uteis, axis=1, df=df_timeline)
        # df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
        #     lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    # df_timeline = df_timeline[['datafim','diferenca']]

    df_timeline['maquina'] = df_timeline['maquina']
    df_timeline['maquina'] = df_timeline['maquina'].str.split(' - ').str[0]

    df_agrupado_tempo = df_timeline.groupby(
        ['maquina'])['diferenca'].sum().reset_index()
    df_agrupado_tempo['maquina'] = df_agrupado_tempo['maquina'].str.split(
        ' - ').str[0]

    df_agrupado_qtd = df_timeline[['maquina']]

    # Contar a quantidade de manutenções por máquina
    contagem = df_agrupado_qtd['maquina'].value_counts()
    df_agrupado_qtd['qtd_manutencao'] = df_agrupado_qtd['maquina'].map(
        contagem)
    df_agrupado_qtd = df_agrupado_qtd.drop_duplicates()

    df_combinado = df_agrupado_qtd.merge(df_agrupado_tempo, on='maquina')

    s = ("""
    SELECT * FROM tb_maquinas
    """)

    df_maquinas = pd.read_sql_query(s, conn).drop_duplicates()
    df_maquinas = df_maquinas[['codigo', 'apelido']]
    df_maquinas = df_maquinas.rename(columns={'codigo': 'maquina'})

    df_combinado = df_combinado.merge(
        df_maquinas, left_on='maquina', right_on='apelido')
    df_combinado['diferenca'] = df_combinado['diferenca'] / 60
    df_combinado = df_combinado.drop_duplicates()

    df_combinado = df_combinado.reset_index(drop=True)

    for i in range(len(df_combinado)):
        if df_combinado['apelido'][i] == '':
            df_combinado['apelido'][i] = df_combinado['maquina_y'][i]

    qtd_dias_uteis = dias_uteis_inicial_final(dia_inicial,dia_final)

    df_combinado['carga_trabalhada'] = qtd_dias_uteis * 9

    df_combinado['MTTR'] = df_combinado['diferenca'] / df_combinado['qtd_manutencao']

    # df_combinado['diferenca'] = df_combinado['diferenca']

    # df_combinado['MTTR'] = df_combinado['MTTR']

    df_combinado_mttr = df_combinado[[
        'apelido', 'qtd_manutencao', 'diferenca', 'MTTR']].values.tolist()

    if len(df_combinado) > 0:

        df_combinado['diferenca'] = df_combinado['diferenca'].round(2)
        df_combinado['MTTR'] = df_combinado['MTTR'].round(2)

        df_combinado.sort_values("MTTR", inplace=True)

        grafico1_maquina = df_combinado['apelido'].tolist()  # eixo x
        grafico2_mttr = df_combinado['MTTR'].tolist()  # eixo y grafico 2

        # sorted_tuples = sorted(zip(grafico1_maquina, grafico2_mttr), key=lambda x: x[0])

        # # Desempacotar as tuplas classificadas em duas listas separadas
        # grafico1_maquina, grafico2_mttr = zip(*sorted_tuples)

        # grafico1_maquina = list(grafico1_maquina)
        # grafico2_mttr = list(grafico2_mttr)

        context_mttr_maquina = {
            'labels_mttr_maquina': grafico1_maquina, 'dados_mttr_maquina': grafico2_mttr}

    else:

        grafico1_maquina = []
        grafico2_mttr = []

        context_mttr_maquina = {
            'labels_mttr_maquina': grafico1_maquina, 'dados_mttr_maquina': grafico2_mttr}

    """
    Gráfico de MTTR por setor
    utiliza a query_mttr
    """

    # query_mttr
    # mttr_setor

    df_timeline = pd.read_sql_query(query_mttr, conn)

    df_timeline = df_timeline.sort_values(by='n_ordem')

    df_timeline = df_timeline.dropna()

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    df_timeline = df_timeline.dropna()

    df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])
    # Crie uma nova coluna 'mes' usando o atributo .dt.month
    df_timeline['mes'] = df_timeline['fim'].dt.month

    df_timeline['now'] = datetime.now()
    df_timeline['mes_hoje'] = datetime.now().month

    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
            lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    df_agrupado_tempo = df_timeline.groupby(
        ['setor'])['diferenca'].sum().reset_index()

    df_agrupado_qtd = df_timeline[['setor']]

    # Contar a quantidade de manutenções por máquina
    contagem = df_agrupado_qtd['setor'].value_counts()
    df_agrupado_qtd['qtd_manutencao'] = df_agrupado_qtd['setor'].map(contagem)
    df_agrupado_qtd = df_agrupado_qtd.drop_duplicates()

    df_combinado = df_agrupado_qtd.merge(df_agrupado_tempo, on='setor')

    df_combinado['diferenca'] = df_combinado['diferenca'] / 60

    qtd_dias_uteis = dias_uteis_inicial_final(dia_inicial,dia_final)

    df_combinado['carga_trabalhada'] = qtd_dias_uteis * 9

    df_combinado['MTTR'] = (df_combinado['diferenca'] /
                            df_combinado['qtd_manutencao'])
    df_combinado['diferenca'] = df_combinado['diferenca']

    df_combinado_mttr_setor = df_combinado[[
        'setor', 'qtd_manutencao', 'diferenca', 'MTTR']].values.tolist()

    if len(df_combinado) > 0:

        df_combinado['MTTR'] = (
            df_combinado['diferenca'] / df_combinado['qtd_manutencao']).round(2)
        df_combinado['diferenca'] = df_combinado['diferenca'].round(2)

        df_combinado.sort_values("MTTR", inplace=True)

        grafico1_maquina = df_combinado['setor'].tolist()  # eixo x
        grafico2_mttr = df_combinado['MTTR'].tolist()  # eixo y grafico 2

        # sorted_tuples = sorted(zip(grafico1_maquina, grafico2_mttr), key=lambda x: x[0])

        # # Desempacotar as tuplas classificadas em duas listas separadas
        # grafico1_maquina, grafico2_mttr = zip(*sorted_tuples)

        # grafico1_maquina = list(grafico1_maquina)
        # grafico2_mttr = list(grafico2_mttr)

        context_mttr_setor = {
            'labels_mttr_setor': grafico1_maquina, 'dados_mttr_setor': grafico2_mttr}

    else:

        grafico1_maquina = []
        grafico2_mttr = []

        context_mttr_setor = {
            'labels_mttr_setor': grafico1_maquina, 'dados_mttr_setor': grafico2_mttr}

    """
    Gráfico de DISPONIBILIDADE por máquina
    utiliza a query_disponibilidade
    """

    # query_disponibilidade 
    # calculo_indicadores_disponibilidade_maquinas

    df_timeline = pd.read_sql_query(query_disponibilidade, conn)

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    df_timeline = df_timeline.dropna()

    df_timeline['datafim'] = pd.to_datetime(df_timeline['datafim'])
    df_timeline['mes'] = df_timeline['datafim'].dt.month

    df_timeline['now'] = datetime.now()
    df_timeline['mes_hoje'] = datetime.now().month
    
    # df_timeline = df_timeline.drop_duplicates(subset=['id_ordem','n_ordem'])

    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = df_timeline.apply(calcular_minutos_uteis, axis=1, df=df_timeline)
        # df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
        #     lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    # df_timeline = df_timeline[['datafim','diferenca']]

    df_timeline['maquina'] = df_timeline['maquina']
    df_timeline['maquina'] = df_timeline['maquina'].str.split(' - ').str[0]

    # df_timeline[df_timeline['maquina'] == 'ES-DESBOBI']

    df_agrupado_tempo = df_timeline.groupby(
        ['maquina'])['diferenca'].sum().reset_index()
    df_agrupado_tempo['maquina'] = df_agrupado_tempo['maquina'].str.split(
        ' - ').str[0]

    df_agrupado_qtd = df_timeline[['maquina']]

    # Contar a quantidade de manutenções por máquina
    contagem = df_agrupado_qtd['maquina'].value_counts()
    df_agrupado_qtd['qtd_manutencao'] = df_agrupado_qtd['maquina'].map(
        contagem)
    df_agrupado_qtd = df_agrupado_qtd.drop_duplicates()

    df_combinado = df_agrupado_qtd.merge(df_agrupado_tempo, on='maquina')

    s = ("""
    SELECT * FROM tb_maquinas
    """)

    df_maquinas = pd.read_sql_query(s, conn).drop_duplicates()
    df_maquinas = df_maquinas[['codigo']]
    df_maquinas = df_maquinas.rename(columns={'codigo': 'maquina'})

    df_combinado = df_combinado.merge(df_maquinas, on='maquina')
    df_combinado['diferenca'] = df_combinado['diferenca'] / 60

    qtd_dias_uteis = dias_uteis(lista_meses)

    df_combinado['carga_trabalhada'] = qtd_dias_uteis * 9

    df_combinado = df_combinado.drop_duplicates().reset_index(drop=True)

    if len(df_combinado) > 0:
        print('Entrou')
        df_combinado['MTBF'] = ((df_combinado['carga_trabalhada'] -df_combinado['diferenca']) / df_combinado['qtd_manutencao']).round(2)
        df_combinado['MTTR'] = (df_combinado['diferenca'] / df_combinado['qtd_manutencao']).round(2)
        df_combinado['disponibilidade'] = ((df_combinado['MTBF'] / (df_combinado['MTBF'] + df_combinado['MTTR'])) * 100).round(2)
        disponibilidade_geral_maquina = df_combinado['disponibilidade'].mean().round(2)

        if boleano_historico == True:
            """
            Se for GET pega todo o histórico e adiciona no atual
            """
            print('Entrou no Boleano')
            historico_csv = pd.read_csv(
                "disponibilidade_historico.csv", sep=';')

            if setor_selecionado:
                historico_csv = historico_csv[historico_csv['setor']
                                              == setor_selecionado]

            historico_csv['maquina'] = historico_csv['maquina'].str.split(
                ' - ').str[0]
            historico_csv['disponibilidade_historico_media'] = historico_csv['disponibilidade_historico_media'].str.replace(
                ',', '.').str.replace("%", "").astype(float)
            df_combinado_disponibilidade = df_combinado.merge(
                historico_csv, how='outer', on='maquina').fillna(100)
            df_combinado_disponibilidade["disponibilidade_media"] = (
                df_combinado_disponibilidade["disponibilidade"] + df_combinado_disponibilidade["disponibilidade_historico_media"]) / 2
            df_combinado_disponibilidade = df_combinado_disponibilidade.drop(columns={
                                                                             "disponibilidade"})
            df_combinado_disponibilidade = df_combinado_disponibilidade.rename(
                columns={"disponibilidade_media": 'disponibilidade'})

            df_combinado_disponibilidade.sort_values(
                by='disponibilidade', inplace=True)

            labels = df_combinado_disponibilidade['maquina'].tolist()  # eixo x
            # eixo y gráfico 1
            dados_disponibilidade = df_combinado_disponibilidade['disponibilidade'].tolist(
            )

            disponibilidade_geral_maquina = df_combinado_disponibilidade['disponibilidade'].mean(
            )

            df_combinado_disponibilidade = df_combinado_disponibilidade[[
                'maquina', 'MTBF', 'MTTR', 'disponibilidade']].values.tolist()

        else:
            print('Não Entrou no Boleano')
            labels = df_combinado['maquina'].tolist()  # eixo x
            # eixo y gráfico 1
            dados_disponibilidade = df_combinado['disponibilidade'].tolist()

            df_combinado_disponibilidade = df_combinado[[
                'maquina', 'MTBF', 'MTTR', 'disponibilidade']].values.tolist()

        context_disponibilidade = {'labels_disponibilidade_maquina': labels,
                                   'dados_disponibilidade_maquina': dados_disponibilidade,
                                   'valor_disponibilidade_geral_maquina': disponibilidade_geral_maquina}

    else:
        print('Não Entrou')
        labels = []
        dados_disponibilidade = []
        df_combinado_disponibilidade = []
        disponibilidade_geral_maquina = []

        context_disponibilidade = {'labels_disponibilidade_maquina': labels,
                                   'dados_disponibilidade_maquina': dados_disponibilidade,
                                   'valor_disponibilidade_geral_maquina': disponibilidade_geral_maquina}

    """
    Gráfico de DISPONIBILIDADE por setor
    utiliza a query_disponibilidade
    """

    # query_disponibilidade 
    # calculo_indicadores_disponibilidade_setor

    df_timeline = pd.read_sql_query(query_disponibilidade, conn)

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    df_timeline = df_timeline.dropna()

    df_timeline['datafim'] = pd.to_datetime(df_timeline['datafim'])
    df_timeline['mes'] = df_timeline['datafim'].dt.month
    
    df_timeline['now'] = datetime.now()
    df_timeline['mes_hoje'] = datetime.now().month

    df_timeline['ultimo_dia_mes'] = ultimo_dia_mes(lista_meses)

    df_timeline.dtypes

    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = df_timeline.apply(calcular_minutos_uteis, axis=1, df=df_timeline)
        # df_timeline['diferenca'] = 
        # df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
        #     lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    # df_timeline = df_timeline[['datafim','diferenca']]

    df_timeline['maquina'] = df_timeline['maquina']
    df_timeline['maquina'] = df_timeline['maquina'].str.split(' - ').str[0]

    if df_timeline['maquina'].drop_duplicates().tolist() == ['Outros']:
        df_agrupado_tempo = df_timeline.groupby('maquina').agg({
            'diferenca': 'sum',
            'setor': 'first'  # Ou use a função de agregação desejada para 'setor'
        }).reset_index()
    else:
        df_agrupado_tempo = df_timeline.groupby(
            ['maquina'])['diferenca'].sum().reset_index()
        
    df_agrupado_tempo['maquina'] = df_agrupado_tempo['maquina'].str.split(
        ' - ').str[0]

    df_agrupado_qtd = df_timeline[['maquina']]

    # Contar a quantidade de manutenções por máquina
    contagem = df_agrupado_qtd['maquina'].value_counts()
    df_agrupado_qtd['qtd_manutencao'] = df_agrupado_qtd['maquina'].map(
        contagem)
    df_agrupado_qtd = df_agrupado_qtd.drop_duplicates()

    df_combinado = df_agrupado_qtd.merge(df_agrupado_tempo, on='maquina')

    s = ("""
    SELECT * FROM tb_maquinas
    """)

    df_maquinas = pd.read_sql_query(s, conn).drop_duplicates()
    df_maquinas = df_maquinas[['codigo','setor']]
    df_maquinas = df_maquinas.rename(columns={'codigo': 'maquina'})

    if df_combinado['maquina'].drop_duplicates().tolist() != ['Outros']:
        df_combinado = df_combinado.merge(df_maquinas, on='maquina')

    df_combinado['diferenca'] = df_combinado['diferenca'] / 60

    qtd_dias_uteis = dias_uteis(lista_meses)

    df_combinado['carga_trabalhada'] = qtd_dias_uteis * 9

    df_combinado = df_combinado.drop_duplicates().reset_index(drop=True)

    if len(df_combinado) > 0:
        print('Entrou')
        df_combinado['MTBF'] = ((df_combinado['carga_trabalhada'] -
                                df_combinado['diferenca']) / df_combinado['qtd_manutencao']).round(2)
        df_combinado['MTTR'] = (
            df_combinado['diferenca'] / df_combinado['qtd_manutencao']).round(2)
        df_combinado['disponibilidade'] = (
            (df_combinado['MTBF'] / (df_combinado['MTBF'] + df_combinado['MTTR'])) * 100).round(2)

        disponibilidade_geral_setor = df_combinado['disponibilidade'].mean().round(
            2)
    
        df_disponibilidade_setor = df_combinado[['setor','MTBF','MTTR','disponibilidade']].groupby('setor').mean()[['MTBF','MTTR','disponibilidade']].reset_index()
        df_disponibilidade_setor['MTBF'] = df_disponibilidade_setor['MTBF'].round(2)
        df_disponibilidade_setor['MTTR'] = df_disponibilidade_setor['MTTR'].round(2)
        df_disponibilidade_setor['disponibilidade'] = df_disponibilidade_setor['disponibilidade'].round(2)

        if boleano_historico == True:
            """
            Se for GET pega todo o histórico e adiciona no atual
            """
            print('Entrou no Boleano')
            historico_csv = pd.read_csv("disponibilidade_historico.csv", sep=';')

            if setor_selecionado:
                historico_csv = historico_csv[historico_csv['setor'] == setor_selecionado]

            historico_csv['setor'] = historico_csv['setor'].str.split(' - ').str[0]
            historico_csv['disponibilidade_historico_media'] = historico_csv['disponibilidade_historico_media'].str.replace(',', '.').str.replace("%", "").astype(float)

            df_disponibilidade_setor = df_combinado.merge(historico_csv, how='outer', on='setor').fillna(100)
            df_disponibilidade_setor["disponibilidade_media"] = (df_disponibilidade_setor["disponibilidade"] + df_disponibilidade_setor["disponibilidade_historico_media"]) / 2
            df_disponibilidade_setor = df_disponibilidade_setor.drop(columns={"disponibilidade"})
            df_disponibilidade_setor = df_disponibilidade_setor.rename(columns={"disponibilidade_media": 'disponibilidade'})

            # Calcular a média da disponibilidade para cada setor
            df_disponibilidade_setor = df_disponibilidade_setor.groupby('setor')[['MTBF','MTTR','disponibilidade']].mean().reset_index()

            df_disponibilidade_setor['MTBF'] = df_disponibilidade_setor['MTBF'].round(2)
            df_disponibilidade_setor['MTTR'] = df_disponibilidade_setor['MTTR'].round(2)
            df_disponibilidade_setor['disponibilidade'] = df_disponibilidade_setor['disponibilidade'].round(2)

            print('2',df_disponibilidade_setor)

            df_disponibilidade_setor.sort_values(by='disponibilidade', inplace=True)

            labels = df_disponibilidade_setor['setor'].tolist()  # eixo x
            dados_disponibilidade = df_disponibilidade_setor['disponibilidade'].tolist()

            disponibilidade_geral_setor = df_disponibilidade_setor['disponibilidade'].mean().round(2)

            df_disponibilidade_setor = df_disponibilidade_setor[['setor', 'MTBF', 'MTTR', 'disponibilidade']].values.tolist()
        

        else:
            print('Não Entrou no Boleano')
            labels = df_disponibilidade_setor['setor'].tolist()  # eixo x
            # eixo y gráfico 1
            dados_disponibilidade = df_disponibilidade_setor['disponibilidade'].tolist()

            df_disponibilidade_setor = df_disponibilidade_setor[[
                'setor', 'MTBF', 'MTTR', 'disponibilidade']].values.tolist()

        context_disponibilidade_setor = {
            'labels_disponibilidade_setor': labels,
            'dados_disponibilidade_setor': dados_disponibilidade,
            'valor_disponibilidade_geral_setor': disponibilidade_geral_setor}

    else:
        print('Não Entrou')
        labels = []
        dados_disponibilidade = []
        df_disponibilidade_setor = []
        df_disponibilidade_setor = df_combinado[[
            'setor', 'MTBF', 'MTTR', 'disponibilidade']].values.tolist()
        disponibilidade_geral_setor = []

        context_disponibilidade_setor = {
            'labels_disponibilidade_setor': labels,
            'dados_disponibilidade_setor': dados_disponibilidade,
            'valor_disponibilidade_geral_setor': disponibilidade_geral_setor}

    """
    Gráfico de horas trabalhadas por tipo de manutenção
    utiliza a query_horas_trabalhada_tipo
    """

    # query_horas_trabalhada_tipo
    # horas_trabalhadas_tipo

    df_horas_tipo = pd.read_sql_query(query_horas_trabalhada_tipo, conn)
    # Converter a coluna 'diferenca' para o tipo 'timedelta'
    df_horas_tipo['diferenca'] = pd.to_timedelta(df_horas_tipo['diferenca'])

    # Extrair o total de horas a partir do timedelta e armazenar em uma nova coluna 'diferenca_horas'
    df_horas_tipo['diferenca_horas'] = (
        df_horas_tipo['diferenca'] / pd.Timedelta(hours=1)).round(2)

    # Descartar a coluna 'diferenca' original, se necessário
    df_horas_tipo.drop(columns=['diferenca'], inplace=True)

    df_horas_tipo = df_horas_tipo.dropna()

    lista_horas_trabalhadas_tipo = df_horas_tipo.values.tolist()

    if len(df_horas_tipo) > 0:

        grafico1_maquina = df_horas_tipo['tipo_manutencao'].tolist()  # eixo x
        # eixo y grafico 2
        grafico2_diferenca = df_horas_tipo['diferenca_horas'].tolist()

        sorted_tuples = sorted(
            zip(grafico1_maquina, grafico2_diferenca), key=lambda x: x[0])

        # Desempacotar as tuplas classificadas em duas listas separadas
        grafico1_maquina, grafico2_diferenca = zip(*sorted_tuples)

        grafico1_maquina = list(grafico1_maquina)
        grafico2_diferenca = list(grafico2_diferenca)

        context_horas_trabalhadas_tipo = {
            'labels_horas_trabalhadas_tipo': grafico1_maquina, 'dados_horas_trabalhadas_tipo': grafico2_diferenca}

    else:

        grafico1_maquina = []
        grafico2_diferenca = []

        context_horas_trabalhadas_tipo = {
            'labels_horas_trabalhadas_tipo': grafico1_maquina, 'dados_horas_trabalhadas_tipo': grafico2_diferenca}

    

    """
    Gráfico de horas trabalhadas por área de manutenção
    utiliza a query_horas_trabalhada_area
    """

    # query_horas_trabalhada_area

    df_horas_area = pd.read_sql_query(query_horas_trabalhada_area, conn)
    # Converter a coluna 'diferenca' para o tipo 'timedelta'
    df_horas_area['diferenca'] = pd.to_timedelta(df_horas_area['diferenca'])

    # Extrair o total de horas a partir do timedelta e armazenar em uma nova coluna 'diferenca_horas'
    df_horas_area['diferenca_horas'] = (
        df_horas_area['diferenca'] / pd.Timedelta(hours=1)).round(2)

    # Descartar a coluna 'diferenca' original, se necessário
    df_horas_area.drop(columns=['diferenca'], inplace=True)

    df_horas_area = df_horas_area.dropna()

    lista_horas_trabalhadas_area = df_horas_area.values.tolist()

    if len(df_horas_area) > 0:

        grafico1_maquina = df_horas_area['area_manutencao'].tolist()  # eixo x
        # eixo y grafico 2
        grafico2_diferenca = df_horas_area['diferenca_horas'].tolist()

        sorted_tuples = sorted(
            zip(grafico1_maquina, grafico2_diferenca), key=lambda x: x[0])

        # Desempacotar as tuplas classificadas em duas listas separadas
        grafico1_maquina, grafico2_diferenca = zip(*sorted_tuples)

        grafico1_maquina = list(grafico1_maquina)
        grafico2_diferenca = list(grafico2_diferenca)

        context_horas_trabalhadas_area = {
            'labels_horas_trabalhadas_area': grafico1_maquina, 'dados_horas_trabalhadas_area': grafico2_diferenca}

    else:

        grafico1_maquina = []
        grafico2_diferenca = []

        context_horas_trabalhadas_area = {
            'labels_horas_trabalhadas_area': grafico1_maquina, 'dados_horas_trabalhadas_area': grafico2_diferenca}

    """
    Gráfico de horas trabalhadas por setor
    utiliza a query_horas_trabalhadas_setor
    """

    # query_horas_trabalhadas_setor

    df_horas_tipo = pd.read_sql_query(query_horas_trabalhada_setor, conn)

    # Converter a coluna 'diferenca' para o tipo 'timedelta'
    df_horas_tipo['diferenca'] = pd.to_timedelta(df_horas_tipo['diferenca'])

    # Extrair o total de horas a partir do timedelta e armazenar em uma nova coluna 'diferenca_horas'
    df_horas_tipo['diferenca_horas'] = (
        df_horas_tipo['diferenca'] / pd.Timedelta(hours=1)).round(2)

    # Descartar a coluna 'diferenca' original, se necessário
    df_horas_tipo.drop(columns=['diferenca'], inplace=True)

    df_horas_tipo = df_horas_tipo.dropna()

    lista_horas_trabalhadas_setor = df_horas_tipo.values.tolist()

    if len(df_horas_tipo) > 0:

        grafico1_maquina = df_horas_tipo['setor'].tolist()  # eixo x
        # eixo y grafico 2
        grafico2_diferenca = df_horas_tipo['diferenca_horas'].tolist()

        sorted_tuples = sorted(
            zip(grafico1_maquina, grafico2_diferenca), key=lambda x: x[0])

        # Desempacotar as tuplas classificadas em duas listas separadas
        grafico1_maquina, grafico2_diferenca = zip(*sorted_tuples)

        grafico1_maquina = list(grafico1_maquina)
        grafico2_diferenca = list(grafico2_diferenca)

        context_horas_trabalhadas_setor = {
            'labels_horas_trabalhadas_setor': grafico1_maquina, 'dados_horas_trabalhadas_setor': grafico2_diferenca}

    else:

        grafico1_maquina = []
        grafico2_diferenca = []

        context_horas_trabalhadas_setor = {
            'labels_horas_trabalhadas_setor': grafico1_maquina, 'dados_horas_trabalhadas_setor': grafico2_diferenca}

    """
    Gráfico de top 10 MTBF por máquina
    utiliza a query_mtbf
    """

    # query mtbf
    # top 10

    df_timeline = pd.read_sql_query(query_mtbf, conn)

    df_timeline['datafim'] = pd.to_datetime(df_timeline['datafim'])
    # Crie uma nova coluna 'mes' usando o atributo .dt.month
    df_timeline['mes'] = df_timeline['datafim'].dt.month

    # mes_hoje = datetime.today().month - 1
    # df_timeline = df_timeline[df_timeline['mes'] == mes_hoje]

    df_timeline = df_timeline.dropna()

    df_timeline['maquina'] = df_timeline['maquina']
    df_timeline['maquina'] = df_timeline['maquina'].str.split(' - ').str[0]

    df_timeline['maquina'].value_counts()

    # Contar a quantidade de manutenções por máquina
    contagem = df_timeline['maquina'].value_counts()
    df_timeline['qtd_manutencao'] = df_timeline['maquina'].map(contagem)
    df_timeline = df_timeline.drop_duplicates(subset='maquina')

    qtd_dias_uteis = dias_uteis_inicial_final(dia_inicial,dia_final)

    df_timeline['carga_trabalhada'] = qtd_dias_uteis * 9

    df_timeline['MTBF'] = ((df_timeline['carga_trabalhada']) /
                           df_timeline['qtd_manutencao']).round(2)
    print(df_timeline)

    # top_10_maiores_MTBF_lista = top_10_maiores_MTBF[['maquina','qtd_manutencao','carga_trabalhada','MTBF']].values.tolist()

    if len(df_timeline) > 0:

        if boleano_historico == True:
            """
            Se for GET pega todo o histórico e adiciona no atual
            """
            historico_csv = pd.read_csv("mtbf_historico.csv", sep=';')
            historico_csv["historico_mtbf_decimal"] = historico_csv["historico_mtbf"].apply(
                convert_time_to_decimal)

            # if setor_selecionado:
            #     historico_csv = historico_csv[historico_csv['setor']
            #                                   == setor_selecionado]

            historico_csv['maquina'] = historico_csv['maquina'].str.split(
                ' - ').str[0]

            df_timeline = df_timeline.merge(
                historico_csv, how='outer', on='maquina').fillna(0)
            df_timeline['MTBF_final'] = df_timeline['historico_mtbf_decimal'] + \
                df_timeline['MTBF']
            df_timeline = df_timeline.drop(columns={'MTBF'})
            df_timeline = df_timeline.rename(
                columns={"MTBF_final": 'MTBF'}).round(2)
            df_timeline = df_timeline.iloc[:10]
            df_timeline.sort_values("MTBF", inplace=True)
            media_qtd_manutencao = df_timeline['qtd_manutencao'][df_timeline['qtd_manutencao'] != 0].mean()

            # Substitui os valores 0 em qtd_manutencao pela média calculada
            df_timeline['qtd_manutencao'] = df_timeline['qtd_manutencao'].replace(0,round(media_qtd_manutencao))

            df_timeline['carga_trabalhada'] = qtd_dias_uteis * 9

            grafico1_top10_maquina = df_timeline['maquina'].tolist()  # eixo x
            # eixo y gráfico 1
            grafico1_top10_mtbf = df_timeline['MTBF'].tolist()

            top_10_maiores_MTBF_lista = df_timeline[[
                'maquina', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        else:

            df_timeline = df_timeline.iloc[:10]
            df_timeline.sort_values("MTBF", inplace=True)
            top_10_maiores_MTBF_lista = df_timeline[[
                'maquina', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

            grafico1_top10_maquina = df_timeline['maquina'].tolist()  # eixo x
            # eixo y gráfico 1
            grafico1_top10_mtbf = df_timeline['MTBF'].tolist()

        context_mtbf_top10_maquina = {
            'labels_mtbf_top10_maquina': grafico1_top10_maquina, 'dados_mtbf_top10_maquina': grafico1_top10_mtbf}
    else:

        grafico1_top10_maquina = []
        grafico1_top10_mtbf = []
        top_10_maiores_MTBF_lista = df_timeline[[
            'maquina', 'qtd_manutencao', 'carga_trabalhada', 'MTBF']].values.tolist()

        context_mtbf_top10_maquina = {
            'labels_mtbf_top10_maquina': grafico1_top10_maquina, 'dados_mtbf_top10_maquina': grafico1_top10_mtbf}

    # Organize os resultados em um dicionário
    resultado = {
        'context_mtbf_maquina': context_mtbf_maquina,
        'df_timeline_copia': df_timeline_copia,

        'context_mtbf_setor': context_mtbf_setor,
        'df_timeline_mtbf_setor': df_timeline_mtbf_setor,

        'context_mttr_maquina': context_mttr_maquina,
        'df_combinado_mttr': df_combinado_mttr,

        'context_mttr_setor': context_mttr_setor,
        'df_combinado_mttr_setor': df_combinado_mttr_setor,

        'context_horas_trabalhadas': context_horas_trabalhadas_setor,
        'lista_horas_trabalhadas': lista_horas_trabalhadas_setor,

        'context_disponibilidade': context_disponibilidade,
        'df_combinado_disponibilidade': df_combinado_disponibilidade,
        'disponibilidade_geral_maquina': disponibilidade_geral_maquina,

        'context_disponibilidade_setor': context_disponibilidade_setor,
        'df_disponibilidade_setor': df_disponibilidade_setor,
        'disponibilidade_geral_setor': disponibilidade_geral_setor,

        'context_horas_trabalhadas_tipo': context_horas_trabalhadas_tipo,
        'lista_horas_trabalhadas_tipo': lista_horas_trabalhadas_tipo,

        'context_horas_trabalhadas_area': context_horas_trabalhadas_area,
        'lista_horas_trabalhadas_area': lista_horas_trabalhadas_area,

        'context_mtbf_top10_maquina': context_mtbf_top10_maquina,
        'top_10_maiores_MTBF_lista': top_10_maiores_MTBF_lista,

    }

    return resultado


def formulario_os(id_ordem):

    """
    Função para gerar arquivo excel com informações sobre a OS.
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """SELECT * FROM tb_ordens WHERE id_ordem = {}""".format(id_ordem)
    cur.execute(query)
    df = pd.read_sql_query(query, conn)

    ultima_atualizacao = df['ultima_atualizacao'][0] - timedelta(hours=3)

    wb = load_workbook('modelo_os_new.xlsx')
    ws = wb.active

    nova_hora_formatada = ultima_atualizacao.strftime('%H:%M')
    data_atual = ultima_atualizacao.strftime('%d/%m/%Y')

    ws['G8'] = data_atual
    ws['G9'] = nova_hora_formatada

    ws['B10'] = df['solicitante'][0]
    ws['B8'] = df['id_ordem'][0]
    ws['B9'] = df['setor'][0]
    ws['B11'] = df['maquina'][0]
    ws['B12'] = df['problemaaparente'][0]

    if df['maquina_parada'][0] == True:
        ws['G11'] = 'Sim'
    else:
        ws['G11'] = 'Não'

    df = df.drop_duplicates(subset=['id_ordem'], keep='last').reset_index()

    ws['G10'] = df['status'][0]

    wb.save('modelo_os_new.xlsx')

    # workbook = Workbook("modelo_os_new.xlsx")
    # workbook.save("Ordem de Serviço.pdf")

    # convertapi.api_secret = 'vkVdyOJxS8xz4uWq'
    # convertapi.convert('pdf', {
    #     'File': 'modelo_os_new.xlsx'
    # }, from_format = 'xlsx').save_files('modelo_os_new.pdf')

    # arquivo_final = 'modelo_os_new.pdf'

    # Retorna o arquivo para download
    return send_file("modelo_os_new.xlsx", as_attachment=True)


def mes_atual():

    """
    Função para mostrar mês atual
    """

    mesAtual = datetime.now().month

    return mesAtual


def calcular_dias_uteis(ano, mes):

    """
    Função para calcular dias úteis
    """

    dias_uteis = []

    start_date = pd.Timestamp(year=ano, month=mes, day=1)
    end_date = pd.Timestamp(year=ano, month=mes, day=1) + \
        pd.DateOffset(months=1) - pd.DateOffset(days=1)

    for day in pd.date_range(start_date, end_date):
        if day.weekday() < 5:  # 0 a 4 representam os dias da semana de segunda a sexta-feira
            dias_uteis.append(day)

    dias_uteis = len(dias_uteis)

    return dias_uteis


def custo_MO():

    """
    Cálculo de custo da mão obra por ordem de serviço
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = ("""
        SELECT
            id_ordem,
            dataabertura,
            n_ordem,
            status,
            datainicio,
            datafim,
            STRING_AGG(REGEXP_REPLACE(operador, '[^\d,]', '', 'g'), ', ') AS operador,
            MIN(TO_TIMESTAMP(datainicio || ' ' || horainicio, 'YYYY-MM-DD HH24:MI:SS')) AS inicio,
            MAX(TO_TIMESTAMP(datafim || ' ' || horafim, 'YYYY-MM-DD HH24:MI:SS')) AS fim
        FROM (
            SELECT
                id_ordem,
                dataabertura,
                n_ordem,
                status,
                datainicio,
                datafim,
                operador,
                descmanutencao,
                horainicio,
                horafim
            FROM tb_ordens
            WHERE (ordem_excluida IS NULL OR ordem_excluida = FALSE)
        ) subquery
        GROUP BY id_ordem, dataabertura, n_ordem, status, datainicio, datafim, descmanutencao;
        """)

    df_timeline = pd.read_sql_query(s, conn)
    df_funcionario = pd.read_sql_query("SELECT * FROM tb_funcionario", conn) 

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    for i in range(len(df_timeline)):
        if df_timeline['fim'][i] == 'NaT':
            df_timeline['fim'][i] = 0
            df_timeline['inicio'][i] = 0
        else:
            pass

    df_timeline = df_timeline.replace(np.nan, '-')

    df_timeline['operador'] = df_timeline['operador'].apply(lambda x: ', '.join(re.findall(r'\d+', x)))
    
    df_operadores = df_timeline['operador'].str.split(', ', expand=True).stack().reset_index(level=1, drop=True).reset_index()
    df_operadores.columns = ['index', 'operador']

    # Mesclar os DataFrames usando o índice original
    df_resultado = df_timeline.drop(columns=['operador']).merge(df_operadores, left_index=True, right_on='index', how='inner')
    df_timeline = df_resultado.drop(columns=['index'])
    
    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
            lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    df_timeline = df_timeline.sort_values(by='n_ordem', ascending=True)

    if df_timeline['datainicio'][0] == '-':
        df_timeline['datainicio'][0] = df_timeline['dataabertura'][0]

    df_final = df_timeline

    df_final['operador'] = df_final['operador'].replace('',0).astype(int)
    df_funcionario['matricula'] = df_funcionario['matricula'].astype(int)

    df_timeline = df_final.merge(df_funcionario, left_on='operador', right_on='matricula')

    df_timeline['mesExecucao'] = df_timeline['fim'].dt.month
    df_timeline['anoExecucao'] = df_timeline['fim'].dt.year
    df_timeline['dias_uteis'] = df_timeline.apply(
        lambda row: calcular_dias_uteis(row['anoExecucao'], row['mesExecucao']), axis=1)
    df_timeline['horasTotalMes'] = df_timeline['dias_uteis'] * (9*60)
    df_timeline['salario'] = df_timeline['salario'].replace("-", 0)
    df_timeline['salario'] = df_timeline['salario'].astype(float)
    df_timeline['proporcional'] = (
        df_timeline['salario'] * df_timeline['diferenca']) / df_timeline['horasTotalMes']

    df_groupby = df_timeline[['id_ordem', 'proporcional']].groupby(
        ['id_ordem']).sum().reset_index().round(2)

    df_timeline = df_timeline.drop(columns=[
                                   'mesExecucao', 'anoExecucao', 'dias_uteis', 'horasTotalMes', 'proporcional', 'nome', 'matricula', 'salario'])
    df_timeline = df_timeline.drop_duplicates(subset=['id_ordem'])

    df_final = pd.merge(df_timeline, df_groupby, how='left', on='id_ordem')

    df_final['diferenca'] = df_final['diferenca'].astype(int)

    df_final = df_final[['id_ordem', 'proporcional']]

    return df_final


def tempo_maquina_parada():
    
    """
    Função para calcular tempo de máquina parada
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = ("""
        SELECT
            id_ordem,
            dataabertura,
            n_ordem,
            status,
            datainicio,
            datafim,
            STRING_AGG(REGEXP_REPLACE(operador, '[^\d,]', '', 'g'), ', ') AS operador,
            MIN(TO_TIMESTAMP(datainicio || ' ' || horainicio, 'YYYY-MM-DD HH24:MI:SS')) AS inicio,
            MAX(TO_TIMESTAMP(datafim || ' ' || horafim, 'YYYY-MM-DD HH24:MI:SS')) AS fim
        FROM (
            SELECT
                id_ordem,
                dataabertura,
                n_ordem,
                status,
                datainicio,
                datafim,
                operador,
                descmanutencao,
                horainicio,
                horafim
            FROM tb_ordens
            WHERE (ordem_excluida IS NULL OR ordem_excluida = FALSE)
        ) subquery
        GROUP BY id_ordem, dataabertura, n_ordem, status, datainicio, datafim, descmanutencao;
        """)

    df_timeline = pd.read_sql_query(s, conn)

    df_timeline['inicio'] = df_timeline['inicio'].astype(str)
    df_timeline['fim'] = df_timeline['fim'].astype(str)

    for i in range(len(df_timeline)):
        if df_timeline['fim'][i] == 'NaT':
            df_timeline['fim'][i] = 0
            df_timeline['inicio'][i] = 0
        else:
            pass

    df_timeline = df_timeline.replace(np.nan, '-')
        
    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
            lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

    except:
        df_timeline['diferenca'] = 0

    df_timeline = df_timeline.sort_values(by='n_ordem', ascending=True)

    if df_timeline['datainicio'][0] == '-':
        df_timeline['datainicio'][0] = df_timeline['dataabertura'][0]

    df_final = df_timeline

    df_final['operador'] = df_final['operador'].replace('',0).astype(int)


    df_final = df_final[['id_ordem', 'proporcional']]

    return df_final


def allowed_file(filename):

    """
    Função para verificar a extensão do arquivo permitida
    """

    # Lista de extensões permitidas para vídeos
    ALLOWED_EXTENSIONS = {'mp4', 'avi', 'mov'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@routes_bp.route('/')
@login_required
def inicio():  # Redirecionar para a página de login

    """
    Rota para página de login
    """

    return render_template("login/login.html")


@routes_bp.route('/index')
@login_required
def Index():  # Página inicial (Página com a lista de ordens de serviço)

    """
    Rota para página principal da aplicação, mostrando a tabela principal.
    """

    setor_selecionado = session.get('setor')
    identificador_selecionado = session.get('identificador')

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = (""" select DISTINCT t7.*, t8.id_ordem as contem_imagem
            from (
                select t5.*, t6.id_ordem as contem_video
                from(
                    select t3.*, t4.parada1,t4.parada2,t4.parada3
                    from(
                        SELECT DISTINCT t1.total, t2.* 
                        FROM (
                            SELECT tb_carrinho.id_ordem, SUM(tb_material.valor * tb_carrinho.quantidade) AS total
                            FROM tb_carrinho
                            JOIN tb_material ON tb_carrinho.codigo = tb_material.codigo
                            GROUP BY tb_carrinho.id_ordem
                            ) t1
                        RIGHT JOIN tb_ordens t2 ON t1.id_ordem = t2.id_ordem
                    ) as t3
                    LEFT JOIN tb_paradas t4 ON t3.id_ordem = t4.id_ordem
                    ORDER BY t3.id_ordem
                ) as t5
                LEFT JOIN tb_videos_ordem_servico t6 on t5.id_ordem = t6.id_ordem
                ) as t7
            LEFT JOIN tb_imagens t8 on t7.id_ordem = t8.id_ordem;
         """)

    df = pd.read_sql_query(s, conn)
    df = df.sort_values(by='id_ordem').reset_index(drop=True)

    df = df[df['ordem_excluida'] != True].reset_index(drop=True)

    df.fillna('', inplace=True)

    for i in range(len(df)-1, 0, -1):
        if df['id_ordem'][i] == df['id_ordem'][i-1]:
            if df['maquina_parada'][i-1] == '':
                df['maquina_parada'][i-1] = df['maquina_parada'][i]

    for i in range(1, len(df)):
        if df['id_ordem'][i-1] == df['id_ordem'][i]:
            df['maquina_parada'][i] = df['maquina_parada'][i-1]

    df = df.sort_values(by='n_ordem')

    df.reset_index(drop=True, inplace=True)
    df.replace(np.nan, '', inplace=True)

    df['dataabertura'] = df['dataabertura'].fillna(method='ffill')
    df['dataabertura'] = df['dataabertura'].replace('', method='ffill')

    df = df.drop_duplicates(subset=['id_ordem'], keep='last')
    df = df.sort_values(by='id_ordem')
    df.reset_index(drop=True, inplace=True)

    for i in range(len(df)):
        if df['total'][i] == '':
            df['total'][i] = 0

    df['total'] = df['total'].apply(lambda x: round(x, 2))

    df = df.sort_values('ultima_atualizacao', ascending=False)

    df['ultima_atualizacao'] = pd.to_datetime(df['ultima_atualizacao'])
    df['ultima_atualizacao'] = df['ultima_atualizacao'] - timedelta(hours=3)
    df['ultima_atualizacao'] = df['ultima_atualizacao'].dt.strftime(
        "%Y-%m-%d %H:%M:%S")
    
    # .dt.strftime("%d/%m/%Y")

    df.reset_index(drop=True, inplace=True)

    for i in range(len(df)):
        try:
            if df['dataabertura'][i].strftime('%H:%M') == '03:00':
                df['dataabertura'][i] = df['ultima_atualizacao'][i]
        except:
            pass
        
    for i in range(len(df)):
        if df['maquina_parada'][i] == '':
            df['maquina_parada'][i] = False

    for i in range(len(df)):
        if df['status'][i] == 'Finalizada' or df['parada1'][i] == 'false':
            df['maquina_parada'][i] = False

    df_custos = custo_MO()

    df = pd.merge(df, df_custos, how='left', on='id_ordem')

    df['proporcional'] = df['proporcional'].fillna(0)
    list_users = df.values.tolist()

    funcionarios = buscar_funcionarios()

    return render_template('user/index.html', funcionarios=funcionarios, list_users=list_users,setor_selecionado=setor_selecionado,identificador_selecionado=identificador_selecionado)

def proxima_os():
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("SELECT MAX(id_ordem) FROM tb_ordens")
    ultima_os = cur.fetchone()[0]
    proxima_os = ultima_os+1

    return proxima_os

def salvar_imagem(imagens, os):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if not any(imagens):
        return 'sem video'
    else:
        for imagem in imagens:

            # Ler os dados da imagem
            imagem_data = imagem.read()

            # Abrir a imagem usando a biblioteca Pillow
            image = Image.open(io.BytesIO(imagem_data))

            # Converter a imagem para o modo RGB, se necessário
            if image.mode != 'RGB':
                image = image.convert('RGB')

            # Redimensionar a imagem para um tamanho desejado
            max_size = (800, 600)
            image.thumbnail(max_size)

            # Salvar a imagem com uma qualidade reduzida
            buffer = io.BytesIO()
            image.save(buffer, format='JPEG', quality=80)
            imagem_data_comprimida = buffer.getvalue()

            cur = conn.cursor(
                cursor_factory=psycopg2.extras.DictCursor)
            cur.execute("INSERT INTO tb_imagens (id_ordem, imagem) VALUES (%s,%s)",
                        (os, imagem_data_comprimida))
            conn.commit()
    

    return 'sucesso'

def salvar_video(videos, os):
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    for video in videos:
        if video.filename != '':
            # Verificar a extensão do arquivo (opcional)
            if allowed_file(video.filename):
                filename = secure_filename(video.filename)
                # video.save(os.path.join(routes_bp.config['UPLOAD_FOLDER'], filename))

                # Ler os dados do vídeo como um objeto bytes
                video_data = video.read()

                # Inserir os dados do vídeo no banco de dados
                cur.execute(
                    "INSERT INTO tb_videos_ordem_servico (id_ordem, video) VALUES (%s, %s)", (os, video_data))
                conn.commit()

    return 'sucess'

@routes_bp.route('/abrir-os', methods=['POST'])
def abrir_os():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                    password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    setor = request.form.get('inputSetor')
    maquina = request.form.get('inputMaquinaLocal')
    problema = request.form.get('problemaAparente')
    solicitante = request.form.get('inputSolicitante')
    dataAbertura = datetime.now()
    equipamento_em_falha = request.form.get('inputEquipamentoFalha')
    setor_maquina_solda = request.form.get('inputcampoESetorMaquinaSolda')
    qual_ferramenta = request.form.get('inputFerramenta')
    cod_equipamento = request.form.get('inputCodigoFerramenta')
    risco = request.form['inputImpacto']
    n_ordem = 0
    status = 'Em espera'

    if 'checkboxMaquinaParada' in request.form:
        # O checkbox foi marcado
        maquina_parada = True
    else:
        # O checkbox não foi marcado
        maquina_parada = False

    # Receber e salvar imagem
    imagens = request.files.getlist('imagens')
    salvar_imagem(imagens, proxima_os())

    # Receber e salvar vídeo
    videos = request.files.getlist('video')
    salvar_video(videos, proxima_os())
    
    cur.execute("INSERT INTO tb_ordens (setor, maquina, risco,status, problemaaparente, id_ordem, n_ordem ,dataabertura, maquina_parada,solicitante,equipamento_em_falha,setor_maquina_solda,qual_ferramenta, cod_equipamento) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    (setor, maquina, risco, status, problema, proxima_os(), n_ordem, dataAbertura, maquina_parada, solicitante, equipamento_em_falha, setor_maquina_solda, qual_ferramenta, cod_equipamento))
    
    conn.commit()

    return redirect(url_for('routes.open_os'))


@routes_bp.route('/edit/<id_ordem>/<identificador_selecionado>/<setor_selecionado>', methods=['POST', 'GET'])
@login_required
# Página para edição da ordem de serviço (Informar o andamento da ordem)
def get_employee(id_ordem, identificador_selecionado, setor_selecionado):

    """
    Função para criar uma execução para ordem de serviço
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    s = ('SELECT tb_ordens.*, tb_maquinas.tombamento FROM tb_ordens LEFT JOIN tb_maquinas ON tb_ordens.maquina = tb_maquinas.codigo WHERE tb_ordens.id_ordem = {};'.format(int(id_ordem)))
    cur.execute(s)
    data1 = pd.read_sql_query(s, conn)

    data1 = data1.sort_values(by='n_ordem')
    data1.reset_index(drop=True, inplace=True)
    data1.replace(np.nan, '', inplace=True)

    data1.iloc[:,8:]

    try:
        dataabertura = data1['dataabertura'][0] - timedelta(hours=3)
        dataabertura = dataabertura.tz_convert(None).strftime('%Y-%m-%d %H:%M')
    except:
        dataabertura = data1['ultima_atualizacao'][0] - timedelta(hours=3)
        dataabertura = dataabertura.tz_convert(None).strftime('%Y-%m-%d %H:%M')

    # Loop para percorrer todas as linhas da coluna
    for i in range(1, len(data1['dataabertura'])):
        if data1['dataabertura'][i] == '':
            data1['dataabertura'][i] = data1['dataabertura'][i-1]
        if data1['solicitante'][i] == '':
            data1['solicitante'][i] = data1['solicitante'][i-1]
        if data1['equipamento_em_falha'][i] == '':
            data1['equipamento_em_falha'][i] = data1['equipamento_em_falha'][i-1]
        if data1['setor_maquina_solda'][i] == '':
            data1['setor_maquina_solda'][i] = data1['setor_maquina_solda'][i-1]
        if data1['qual_ferramenta'][i] == '':
            data1['qual_ferramenta'][i] = data1['qual_ferramenta'][i-1]
        if data1['cod_equipamento'][i] == '':
            data1['cod_equipamento'][i] = data1['cod_equipamento'][i-1]
        if data1['pvlye'][i] == '':
            data1['pvlye'][i] = data1['pvlye'][i-1]
        if data1['pa_plus'][i] == '':
            data1['pa_plus'][i] = data1['pa_plus'][i-1]
        if data1['tratamento'][i] == '':
            data1['tratamento'][i] = data1['tratamento'][i-1]
        if data1['ph_agua'][i] == '':
            data1['ph_agua'][i] = data1['ph_agua'][i-1]

    data1 = data1.drop_duplicates(subset=['id_ordem'], keep='last')
    data1 = data1.sort_values(by='id_ordem')

    tipo_manutencao = data1['tipo_manutencao'].values.tolist()[0]
    area_manutencao = data1['area_manutencao'].values.tolist()[0]

    pvlye = data1['pvlye'].values.tolist()[0]
    pa_plus = data1['pa_plus'].values.tolist()[0]
    tratamento = data1['tratamento'].values.tolist()[0]
    ph_agua = data1['ph_agua'].values.tolist()[0]

    data1 = data1.values.tolist()
 
    opcaoAtual = data1[0][4]

    lista_opcoes = ['Em execução', 'Finalizada', 'Aguardando material']

    opcoes = []
    opcoes.append(opcaoAtual)

    for opcao in lista_opcoes:
        opcoes.append(opcao)

    opcoes = list(set(opcoes))
    opcoes.remove(opcaoAtual)  # Remove o elemento 'c' da lista
    opcoes.insert(0, opcaoAtual)

    query = """SELECT * FROM tb_funcionario"""
    tb_funcionarios = pd.read_sql_query(query, conn)
    tb_funcionarios['matricula_nome'] = tb_funcionarios['matricula'] + \
        " - " + tb_funcionarios['nome']
    tb_funcionarios = tb_funcionarios[['matricula_nome']].values.tolist()

    query = """SELECT DISTINCT CONCAT(codigo, ' - ', descricao) AS codigo_descricao
            FROM tb_ordens AS t1
            JOIN tb_maquinas AS t2 ON t1.maquina = t2.codigo
            WHERE t1.id_ordem = {}""".format((int(id_ordem)))
    
    cur.execute(query)

    maquinas = cur.fetchall()

    if len(maquinas) == 0:
        maquinas.append('Outros')
    else:
        maquinas = maquinas[0]

    query_maquinas_preventivas = f""" SELECT DISTINCT codigo
                                        FROM tb_ordens AS t1
                                        JOIN tb_maquinas_preventivas AS t2 ON t1.maquina = t2.codigo
                                        WHERE t1.id_ordem = {int(id_ordem)} """
    
    cur.execute(query_maquinas_preventivas)

    maquinas_preventivas = cur.fetchall()

    preventiva = False

    if len(maquinas_preventivas) > 0:

        preventiva = True

    return render_template('user/edit.html', dataabertura=dataabertura, ordem=data1[0], tb_funcionarios=tb_funcionarios,
                           opcoes=opcoes, tipo_manutencao=tipo_manutencao,
                           area_manutencao=area_manutencao, maquinas=maquinas, pvlye=pvlye, pa_plus=pa_plus,
                           tratamento=tratamento, ph_agua=ph_agua,identificador_selecionado=identificador_selecionado,
                           setor_selecionado=setor_selecionado,preventiva=preventiva)

@routes_bp.route('/dados-ordem-servico', methods=['POST'])
@login_required
def dados_ordem_servico():
    data = request.get_json()
    
    data = buscar_dados_os(data['id_ordem'])
    
    return jsonify(data)

@routes_bp.route('/dados-editar-ordem', methods=['POST'])
@login_required
def dados_editar_ordem():
    data = request.get_json()
    print(data)
    data = dados_para_editar(data['id_ordem'], data['n_ordem'])
    
    print(data)

    return jsonify(data)

def verificar_maquina_preventiva(maquina):
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    s = """select codigo from tb_maquinas_preventivas where codigo = %s"""
    cur.execute(s,(maquina,))

    data = cur.fetchall()

    if len(data) > 0:
        return True
    else:
        return False

@routes_bp.route('/guardar-ordem-editada', methods=['POST'])
@login_required
def editar_ordem_banco():
    
    dados = request.get_json()
    print(dados)
    
    # setor = dados['setor']
    maquina = dados['maquina_edicao']
    risco = dados['inputRisco_edicao']
    status = dados['statusLista_edicao']
    # problema = dados['problema']
    id_ordem = dados['numeroOs']
    n_execucao = dados['n_ordem_edicao']
    descmanutencao = dados['descmanutencao_edicao']
    operador = dados['operador_edicao']
    inputEquipamentoEmFalha_edicao = dados['inputEquipamentoEmFalha_edicao']
    codigoEquipamento_edicao = dados['codigoEquipamento_edicao']
    setorMaqSolda_edicao = dados['setorMaqSolda_edicao']
    qual_ferramenta_edicao = dados['qual_ferramenta_edicao']

    matriculas_operadores = []

    # Itere sobre a lista
    for item in operador:
        # Divida a string usando '-' como separador e pegue a parte antes do hífen
        partes = item.split('-')
        if len(partes) > 0:
            matriculas = partes[0].strip()
            matriculas_operadores.append(matriculas)

    operador = matriculas_operadores

    operador = ",".join(operador)

    tipo_manutencao = dados['selectTipoManutencao_edicao']
    datetimes = dados['data_edit_edicao']
    area_manutencao = dados['areaManutencao_edicao']
    pvlye = dados['pvlye_edicao']
    pa_plus = dados['pa-plus_edicao']
    tratamento = dados['tratamento_edicao']
    ph_agua = dados['ph-agua_edicao']

    dataAbertura,natureza = buscar_data_abertura_natureza(id_ordem)

    datainicio,horainicio,datafim,horafim = formatar_data_hora(datetimes)

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute("""update tb_ordens
    set maquina=%s, risco=%s, status=%s, datainicio=%s, horainicio=%s,
        datafim=%s, horafim=%s, descmanutencao=%s,
        operador=%s, tipo_manutencao=%s, area_manutencao=%s,
        equipamento_em_falha=%s,setor_maquina_solda=%s,qual_ferramenta=%s,
        cod_equipamento=%s,pvlye=%s, pa_plus=%s, tratamento=%s, ph_agua=%s
        where id_ordem = %s and n_ordem = %s""", (maquina,risco,status,datainicio,horainicio,
                                                  datafim,horafim,descmanutencao,operador,
                                                  tipo_manutencao,area_manutencao,inputEquipamentoEmFalha_edicao,
                                                  setorMaqSolda_edicao,qual_ferramenta_edicao,codigoEquipamento_edicao,
                                                  pvlye,pa_plus,tratamento,ph_agua,id_ordem,n_execucao))

    conn.commit()

    return 'sucess'

@routes_bp.route('/envio_ok', methods=['POST'])
@login_required
def envio_ok():  # Inserir as edições no banco de dados

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    json_confirmacao = request.get_json()
    
    numero_execucao = int(json_confirmacao['numero_execucao'])

    numeroOrdemValue = int(json_confirmacao['numeroOrdemValue'])

    status = 'Finalizada'

    confirmacao = True

    print(status,confirmacao,numero_execucao,numeroOrdemValue)

    query = "SELECT n_execucao FROM tb_confirmacao;"

    cur.execute(query)

    maquinas_preventivas = cur.fetchall()

    for maquina in maquinas_preventivas:
        print(maquina[0], numeroOrdemValue)
        if maquina[0] == numeroOrdemValue:
            print("Entrou")
            return "Número da Ordem já enviado"

    cur.execute("INSERT INTO tb_confirmacao (n_ordem,n_execucao,confirmacao) VALUES (%s,%s,%s)",(numero_execucao,numeroOrdemValue,confirmacao))

    cur.execute("""
        UPDATE tb_ordens
        SET status=%s, confirmacao=%s
        WHERE n_ordem = %s and id_ordem = %s
        """, (status, confirmacao , numero_execucao - 1, numeroOrdemValue))

    conn.commit()

    return "Sucesso"

@routes_bp.route('/update/<id_ordem>/<identificador_selecionado>/<setor_selecionado>', methods=['POST'])
@login_required
def update_student(id_ordem,identificador_selecionado,setor_selecionado):  # Inserir as edições no banco de dados

    """
    Rota para editar ordem de serviço
    """

    # # Execute a instrução SQL para alterar o tipo da coluna
    # alter_query = "ALTER TABLE tb_ordens ALTER COLUMN tipo_manutencao TYPE TEXT;"
    # cur.execute(alter_query)
    # conn.commit()

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    if request.method == 'POST':

        # cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        # cur.execute(""" 
        #     SELECT MAX(id) FROM tb_ordens
        # """)

        # ultimo_id = cur.fetchone()[0]

        # try:
        #     ultimo_id = ultimo_id+1
        # except:
        #     ultimo_id = 0

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = (""" 
            SELECT natureza FROM tb_ordens where id_ordem = {}
        """).format(id_ordem)

        df = pd.read_sql_query(s, conn)

        natureza = df['natureza'][0]

        setor = request.form['setor']
        maquina = request.form['maquina']
        risco = request.form['risco']
        status = request.form['statusLista']
        problema = request.form['problema']
        id_ordem = id_ordem
        n_ordem = request.form['n_ordem']
        descmanutencao = request.form['descmanutencao']
        operador = request.form.getlist('operador')
        confirmacao = False
        # operador = json.dumps(operador)

        matriculas_operadores = []

        # Itere sobre a lista
        for item in operador:
            # Divida a string usando '-' como separador e pegue a parte antes do hífen
            partes = item.split('-')
            if len(partes) > 0:
                matriculas = partes[0].strip()
                matriculas_operadores.append(matriculas)

        operador = matriculas_operadores

        operador = ",".join(operador)

        print(operador)

        tipo_manutencao = request.form['tipo_manutencao']
        datetimes = request.form['datetimes']
        area_manutencao = request.form['area_manutencao']
        pvlye = request.form.get('pvlye')
        pa_plus = request.form.get('pa-plus')
        tratamento = request.form.get('tratamento')
        ph_agua = request.form.get('ph-agua')

        natureza = natureza

        query_maquinas_preventivas = f""" SELECT DISTINCT codigo
                                        FROM tb_ordens AS t1
                                        JOIN tb_maquinas_preventivas AS t2 ON t1.maquina = t2.codigo
                                        WHERE t1.id_ordem = {int(id_ordem)} """
    
        cur.execute(query_maquinas_preventivas)

        maquinas_preventivas = cur.fetchall()

        if len(maquinas_preventivas) > 0 and status == 'Finalizada':
            status = 'Aguardando OK'
            confirmacao = True
            print("Entrou")
        else:
            print("Entrou no Elsee ")

        try:
            botao1 = request.form['maquina-parada-1']

        except:
            botao1 = 'false'
        try:
            botao2 = request.form['maquina-parada-2']

        except:
            botao2 = 'false'
        try:
            botao3 = request.form['maquina-parada-3']
        except:
            botao3 = 'false'

        if status == 'Finalizada' or status == 'Aguardando OK':
            botao3 = 'true'
            

        print(botao1)
        print(botao2)
        print(botao3)

        # Divida a string em duas partes: data/hora inicial e data/hora final
        data_hora_inicial_str, data_hora_final_str = datetimes.split(" - ")

        # Faça o parsing das strings de data e hora
        data_inicial = datetime.strptime(
            data_hora_inicial_str, "%d/%m/%y %I:%M %p")
        data_final = datetime.strptime(
            data_hora_final_str, "%d/%m/%y %I:%M %p")

        # Formate as datas e horas no formato desejado
        datainicio = data_inicial.strftime("%Y-%m-%d")
        horainicio = data_inicial.strftime("%H:%M:%S")
        datafim = data_final.strftime("%Y-%m-%d")
        horafim = data_final.strftime("%H:%M:%S")

        # print(datainicio, horainicio, datafim, horafim)

        # print(ultimo_id, setor, maquina, risco, status, problema, datainicio, horainicio, datafim, horafim, id_ordem, n_ordem, descmanutencao, [operador])

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute("""
            INSERT INTO tb_ordens ( setor,maquina,risco,status,problemaaparente,
                                    datainicio,horainicio,datafim,horafim,id_ordem,n_ordem,
                                    descmanutencao, operador, natureza, tipo_manutencao,
                                    area_manutencao,pvlye,pa_plus,tratamento,ph_agua,confirmacao) 
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (setor, maquina, risco, status, problema, datainicio, horainicio,
              datafim, horafim, id_ordem, n_ordem, descmanutencao,
              operador, natureza, tipo_manutencao, area_manutencao,
              pvlye, pa_plus, tratamento, ph_agua, confirmacao))

        cur.execute("""
            INSERT INTO tb_paradas (id_ordem,n_ordem, parada1, parada2, parada3) 
                    VALUES (%s,%s,%s,%s,%s)
        """, (id_ordem, n_ordem, botao1, botao2, botao3))

        flash('OS de número {} atualizada com sucesso!'.format(int(id_ordem)))
        conn.commit()

        return redirect(url_for('routes.get_employee', id_ordem=id_ordem,identificador_selecionado=identificador_selecionado,setor_selecionado=setor_selecionado))

def buscar_data_abertura_natureza(ordem_id):
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    query_buscar_data_abertura = "select dataabertura,natureza from tb_ordens where id_ordem = %s and n_ordem = 0"
    
    cur.execute(query_buscar_data_abertura,(ordem_id,))
    data = cur.fetchall()

    dataAbertura = [row['dataabertura'] for row in data][0]
    natureza = [row['natureza'] for row in data][0]

    return dataAbertura,natureza

def formatar_data_hora(datetimes):

    # Divida a string em duas partes: data/hora inicial e data/hora final
    data_hora_inicial_str, data_hora_final_str = datetimes.split(" - ")

    # Faça o parsing das strings de data e hora
    data_inicial = datetime.strptime(
        data_hora_inicial_str, "%d/%m/%Y %H:%M")
    data_final = datetime.strptime(
        data_hora_final_str, "%d/%m/%Y %H:%M")

    # Formate as datas e horas no formato desejado
    datainicio = data_inicial.strftime("%Y-%m-%d")
    horainicio = data_inicial.strftime("%H:%M:%S")
    datafim = data_final.strftime("%Y-%m-%d")
    horafim = data_final.strftime("%H:%M:%S")

    return datainicio,horainicio,datafim,horafim

@routes_bp.route('/executar-ordem', methods=['POST'])
@login_required
def executar_ordem():

    dados = request.get_json()
    
    setor = dados['setor']
    maquina = dados['maquina']
    risco = dados['inputRisco']
    status = dados['statusLista']
    problema = dados['problema']
    id_ordem = dados['numeroOs']
    n_execucao = dados['n_ordem']
    descmanutencao = dados['descmanutencao']
    operador = dados['operador']
    confirmacao = verificar_maquina_preventiva(maquina)
    inputEquipamentoEmFalha = dados['inputEquipamentoEmFalha']
    codigoEquipamento = dados['codigoEquipamento']
    setorMaqSolda = dados['setorMaqSolda']
    qual_ferramenta = dados['qual_ferramenta']
    problemaaparente = dados['problema'] # será usado para descrever o nome do grupo de atividades
    solicitante = dados['inputSolicitante']

    # if status == 'Finalizado' and confirmacao:
    #     confirmacao = True
    #     status = 'Aguardando OK'
    # else:
    #     confirmacao = False
    
    matriculas_operadores = []

    # Itere sobre a lista
    for item in operador:
        # Divida a string usando '-' como separador e pegue a parte antes do hífen
        partes = item.split('-')
        if len(partes) > 0:
            matriculas = partes[0].strip()
            matriculas_operadores.append(matriculas)

    operador = matriculas_operadores

    operador = ",".join(operador)

    tipo_manutencao = dados['selectTipoManutencao']
    datetimes = dados['data_edit']
    area_manutencao = dados['areaManutencao']
    pvlye = dados['pvlye']
    pa_plus = dados['pa-plus']
    tratamento = dados['tratamento']
    ph_agua = dados['ph-agua']

    dataAbertura,natureza = buscar_data_abertura_natureza(id_ordem)

    botao1 = dados['maquina-parada-1']
    botao2 = dados['maquina-parada-2']
    botao3 = dados['maquina-parada-3']
    
    if status == 'Finalizada' or status == 'Aguardando OK':
        botao3 = 'true'
    
    datainicio,horainicio,datafim,horafim = formatar_data_hora(datetimes)

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    print(solicitante, tipo_manutencao, problemaaparente)

    if solicitante == 'Automático' and tipo_manutencao == 'Preventiva':
        cur.execute('update public.tb_grupos_preventivas set ult_manutencao = %s where codigo = %s and grupo = %s', (datafim,maquina,problemaaparente))

    cur.execute("""
        INSERT INTO tb_ordens (setor,maquina,risco,status,problemaaparente,
                                dataabertura,datainicio,horainicio,datafim,horafim,id_ordem,n_ordem,
                                descmanutencao, operador, natureza, tipo_manutencao,
                                area_manutencao,equipamento_em_falha,setor_maquina_solda, 
                                qual_ferramenta,cod_equipamento,
                                pvlye,pa_plus,tratamento,ph_agua,confirmacao) 
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (setor, maquina, risco, status, problema,dataAbertura, datainicio, horainicio,
            datafim, horafim, id_ordem, n_execucao, descmanutencao,
            operador, natureza, tipo_manutencao, area_manutencao,inputEquipamentoEmFalha,
            setorMaqSolda,qual_ferramenta,codigoEquipamento,
            pvlye, pa_plus, tratamento, ph_agua, confirmacao))

    cur.execute("""
        INSERT INTO tb_paradas (id_ordem,n_ordem, parada1, parada2, parada3) 
                VALUES (%s,%s,%s,%s,%s)
    """, (id_ordem, n_execucao, botao1, botao2, botao3))

    conn.commit()

    return 'sucess'

@routes_bp.route('/editar_ordem/<id_ordem>/<n_ordem>', methods=['POST', 'GET'])
@login_required
def editar_ordem(id_ordem, n_ordem):

    """
    Rota para editar execução dentro da ordem de serviço
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = ('SELECT * FROM tb_ordens WHERE id_ordem = {} AND n_ordem = {}'.format(int(id_ordem), int(n_ordem)))
    cur.execute(s)
    data1 = pd.read_sql_query(s, conn)

    data1.reset_index(drop=True, inplace=True)
    data1.replace(np.nan, '', inplace=True)

    desc_manutencao = data1['descmanutencao'].values.tolist()[0]

    executante = data1['operador'].values.tolist()[0].replace(
        "{", "").replace("[", "").replace("\\", "").replace('"', '').replace("]}", "")
    executante = [exec.strip() for exec in executante.split(',')]

    try:
        s = ('SELECT * FROM tb_funcionario')
        cur.execute(s)
        df_funcionarios = pd.read_sql_query(s, conn)

        lista_executante = []

        for matricula in executante:
            funcionario = df_funcionarios[df_funcionarios['matricula'] == matricula][[
                'nome']].values.tolist()[0][0]
            lista_executante.append(matricula + " - " + funcionario)

        executante = lista_executante

    except:
        pass

    tipo_manutencao = data1['tipo_manutencao'].values.tolist()[0]
    area_manutencao = data1['area_manutencao'].values.tolist()[0]

    data_inicio = datetime.strptime(
        str(data1['datainicio'].values[0]), '%Y-%m-%d').strftime('%d/%m/%Y')
    hora_inicio = datetime.strptime(
        str(data1['horainicio'].values[0]), '%H:%M:%S').strftime('%H:%M')
    data_fim = datetime.strptime(
        str(data1['datafim'].values[0]), '%Y-%m-%d').strftime('%d/%m/%Y')
    hora_fim = datetime.strptime(
        str(data1['horafim'].values[0]), '%H:%M:%S').strftime('%H:%M')

    data_atual = f'{data_inicio} {hora_inicio} - {data_fim} {hora_fim}'

    print(data_atual)

    print(data_atual)

    data1 = data1.values.tolist()
    opcaoAtual = data1[0][4]

    cur.close()

    lista_opcoes = ['Em execução', 'Finalizada', 'Aguardando material']

    opcoes = []
    opcoes.append(opcaoAtual)

    for opcao in lista_opcoes:
        opcoes.append(opcao)

    opcoes = list(set(opcoes))
    opcoes.remove(opcaoAtual)  # Remove o elemento 'c' da lista
    opcoes.insert(0, opcaoAtual)

    query = """SELECT * FROM tb_funcionario"""
    tb_funcionarios = pd.read_sql_query(query, conn)
    tb_funcionarios['matricula_nome'] = tb_funcionarios['matricula'] + \
        " - " + tb_funcionarios['nome']
    tb_funcionarios = tb_funcionarios[['matricula_nome']].values.tolist()

    return render_template('user/editar_ordem.html', ordem=data1[0], tb_funcionarios=tb_funcionarios, opcoes=opcoes, tipo_manutencao=tipo_manutencao,
                           area_manutencao=area_manutencao, executante=executante, desc_manutencao=desc_manutencao, data_atual=data_atual)


@routes_bp.route('/editar_ordem_inicial/<id_ordem>/<n_ordem>', methods=['POST', 'GET'])
@login_required
def editar_ordem_inicial(id_ordem, n_ordem):

    """
    Rota para editar a ordem de serviço inicial, por exemplo: data de abertura, máquina, setor...
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    s = ("""SELECT tb_ordens.*, tb_maquinas.tombamento 
        FROM tb_ordens LEFT JOIN tb_maquinas ON tb_ordens.maquina = tb_maquinas.codigo
        WHERE tb_ordens.id_ordem = {} AND tb_ordens.n_ordem = {};""".format(int(id_ordem), int(n_ordem)))

    cur.execute(s)
    data1 = pd.read_sql_query(s, conn)
    
    try:
        data_completa = str(data1['dataabertura'][0] - timedelta(hours=3))
    except:
        data_completa = str(data1['ultima_atualizacao'][0] - timedelta(hours=3))

    try:
        data_datetime = datetime.strptime(data_completa, '%Y-%m-%d %H:%M:%S.%f%z')
    except:
        data_datetime = datetime.strptime(data_completa, '%Y-%m-%d %H:%M:%S+00:00')
    
    if data_datetime.strftime('%H:%M') == '00:00':
        data_completa = str(data1['ultima_atualizacao'][0])
        data_datetime = datetime.strptime(data_completa, '%Y-%m-%d %H:%M:%S.%f%z')

    dataabertura = data_datetime.strftime('%Y-%m-%d %H:%M')

    data1.reset_index(drop=True, inplace=True)
    data1.replace(np.nan, '', inplace=True)
    maquina = data1['maquina'][0]

    try:
        maquina = maquina.split(" - ")[0]
    except:
        pass

    data1 = data1.values.tolist()

    s = ("""
        SELECT CONCAT(codigo , ' - ' , descricao) as codigo_descricao 
        FROM tb_maquinas
        WHERE codigo = '{}'
        """.format(maquina))
    cur.execute(s)

    maquina = cur.fetchall()

    try:
        maquina = maquina[0]
    except:
        maquina = maquina

    if len(maquina) == 0:
        maquina.append('Outros')

    return render_template('user/editar_ordem_inicial.html', ordem=data1[0], maquina=maquina, n_ordem=n_ordem, dataabertura=dataabertura)


@routes_bp.route('/update_ordem/<id_ordem>/<n_ordem>', methods=['POST'])
@login_required
def update_ordem(id_ordem, n_ordem):  # Inserir as edições no banco de dados

    # # Execute a instrução SQL para alterar o tipo da coluna
    # alter_query = "ALTER TABLE tb_ordens ALTER COLUMN tipo_manutencao TYPE TEXT;"
    # cur.execute(alter_query)
    # conn.commit()

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    if request.method == 'POST':

        # cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        # cur.execute(""" 
        #     SELECT MAX(id) FROM tb_ordens
        # """)

        # ultimo_id = cur.fetchone()[0]

        # try:
        #     ultimo_id = ultimo_id+1
        # except:
        #     ultimo_id = 0

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        s = (""" 
            SELECT natureza FROM tb_ordens where id_ordem = {} 
        """).format(id_ordem)

        df = pd.read_sql_query(s, conn)

        natureza = df['natureza'][0]
        # setor = request.form['setor']
        # maquina = request.form['maquina']
        # risco = request.form['risco']
        status = request.form['statusLista']
        # problema = request.form['problema']
        id_ordem = id_ordem
        n_ordem = n_ordem
        descmanutencao = request.form['descmanutencao']
        operador = request.form.getlist('operador')
        operador = json.dumps(operador)
        tipo_manutencao = request.form['tipo_manutencao1']
        datetimes = request.form['datetimes']
        area_manutencao = request.form['area_manutencao1']
        natureza = natureza

        # Divida a string em duas partes: data/hora inicial e data/hora final
        data_hora_inicial_str, data_hora_final_str = datetimes.split(" - ")

        data_inicial = datetime.strptime(
            data_hora_inicial_str, "%d/%m/%Y %H:%M")
        data_final = datetime.strptime(data_hora_final_str, "%d/%m/%Y %H:%M")

        # Formate as datas e horas no formato desejado
        datainicio = data_inicial.strftime("%Y-%m-%d")
        horainicio = data_inicial.strftime("%H:%M:%S")
        datafim = data_final.strftime("%Y-%m-%d")
        horafim = data_final.strftime("%H:%M:%S")

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute("""
        UPDATE tb_ordens
        SET status=%s,datainicio=%s,horainicio=%s,datafim=%s,horafim=%s,id_ordem=%s,
            n_ordem=%s, descmanutencao=%s, operador=%s, natureza=%s, tipo_manutencao=%s, area_manutencao=%s

        WHERE n_ordem = %s and id_ordem = %s
        """, (status, datainicio, horainicio, datafim, horafim, id_ordem, n_ordem, descmanutencao, [operador], natureza, tipo_manutencao, area_manutencao, n_ordem, id_ordem))

        # cur.execute("""
        #     INSERT INTO tb_ordens (id, setor,maquina,risco,status,problemaaparente,datainicio,horainicio,datafim,horafim,id_ordem,n_ordem, descmanutencao, operador, natureza, tipo_manutencao, area_manutencao) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        # """, (ultimo_id, setor, maquina, risco, status, problema, datainicio, horainicio, datafim, horafim, id_ordem, n_ordem, descmanutencao, [operador], natureza, tipo_manutencao, area_manutencao))
        flash('OS de número {} atualizada com sucesso!'.format(int(id_ordem)))
        conn.commit()
        cur.close()

        return redirect(url_for('routes.timeline_os', id_ordem=id_ordem))


@routes_bp.route('/guardar_ordem_editada/<id_ordem>/<n_ordem>', methods=['POST'])
@login_required
# Inserir as edições no banco de dados
def guardar_ordem_editada(id_ordem, n_ordem):

    if request.method == 'POST':

        id_ordem = id_ordem
        setor = request.form['setor']
        maquina = request.form.get('maquina')
        equipamento_em_falha = request.form.get('falha')
        setor_maquina_solda = request.form.get('solda_maquina')
        qual_ferramenta = request.form.get('ferramenta')
        codigo_equipamento = request.form.get("codigo_equip")
        problema = request.form['problema']
        risco = request.form.get("risco")
        maquina_parada = request.form.get('maquina-parada')
        dataabertura = request.form.get('datetimes')

        try:
            maquina = maquina.split(" - ")[0]
        except:
            pass

        if maquina_parada:
            maquina_parada = 'True'
        else:
            maquina_parada = 'False'

        conn = psycopg2.connect(
            dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute("""
            UPDATE tb_ordens
            SET setor=%s,maquina=%s,risco=%s,maquina_parada=%s,equipamento_em_falha=%s,setor_maquina_solda=%s,
            qual_ferramenta=%s,cod_equipamento=%s,problemaaparente=%s,dataabertura=%s
            WHERE id_ordem = %s
            """, (setor, maquina, risco, maquina_parada, equipamento_em_falha, setor_maquina_solda,
                   qual_ferramenta, codigo_equipamento, problema, dataabertura, id_ordem))

        conn.commit()
        cur.close()

        return redirect(url_for('routes.Index'))


def solicitantes():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """SELECT concat(nome_colaborador,' - ',matricula_colaborador) FROM tb_matriculas"""

    cur.execute(query)
    nomes_solicitantes = cur.fetchall()

    return nomes_solicitantes

@routes_bp.route('/openOs')
def open_os():  # Página de abrir OS

    nomes_solicitantes = solicitantes()

    return render_template("user/openOs.html", solicitantes=nomes_solicitantes)


@login_required
@routes_bp.route('/maquinas/<setor>')
def filtro_maquinas(setor):

    # setor = setor.upper()
    if setor == 'Serralheria':
        setor = 'Solda'

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """
        SELECT concat (codigo, ' - ', descricao) FROM tb_maquinas
        WHERE setor = %s
        """

    cur.execute(query,(setor,))
    lista_maquinas = cur.fetchall()
    lista_maquinas.append(["Outros"])

    return jsonify(lista_maquinas)


@routes_bp.route('/edit_material/<id_ordem>', methods=['POST', 'GET'])
@login_required
def get_material(id_ordem):  # Informar material que foi utilizado na ordem de serviço
    # Verifica se a requisição é um POST

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == 'POST':

        # Obtendo o ultimo id

        # cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        # s = ("""
        #     SELECT MAX(id) FROM tb_carrinho
        # """)
        # cur.execute(s)
        # try:
        #     max_id = cur.fetchall()[0][0] + 1
        # except:
        #     max_id = 0

        # Obtém os dados do formulário
        id_ordem = id_ordem
        codigo = request.form['codigo']
        quantidade = request.form['quantidade']

        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
        cur.execute("INSERT INTO tb_carrinho (id_ordem, codigo, quantidade) VALUES (%s,%s,%s)",
                    (id_ordem, codigo, quantidade))
        conn.commit()

    # Obtém os dados da tabela
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    s = ("""
        SELECT tb_carrinho.id_ordem, tb_carrinho.codigo, tb_carrinho.quantidade, tb_material.descricao, tb_material.valor * tb_carrinho.quantidade AS total
        FROM tb_carrinho
        JOIN tb_material ON tb_carrinho.codigo = tb_material.codigo
        WHERE id_ordem = {}
    """).format(int(id_ordem))

    cur.execute(s)
    data = cur.fetchall()

    s = ("""
        SELECT SUM(valortotal.total) AS valor_total FROM
        (
        SELECT tb_carrinho.id_ordem, tb_carrinho.codigo, tb_carrinho.quantidade, tb_material.descricao, SUM(tb_material.valor * tb_carrinho.quantidade) AS total
        FROM tb_carrinho
        JOIN tb_material ON tb_carrinho.codigo = tb_material.codigo
        WHERE tb_carrinho.id_ordem = {}
        GROUP BY tb_carrinho.id_ordem, tb_carrinho.codigo, tb_carrinho.quantidade, tb_material.descricao
        ) AS valortotal; 
    """).format(int(id_ordem))

    cur.execute(s)
    valorTotal = cur.fetchall()

    # cur.close()

    return render_template('user/material.html', datas=data, id_ordem=id_ordem, valorTotal=valorTotal[0][0])


@routes_bp.route('/grafico', methods=['POST', 'GET'])
@login_required
def grafico():  # Dashboard

    if request.method == 'POST':

        boleano_historico = True
        todos_meses = None

        conn = psycopg2.connect(
            dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        cur.execute(
            "SELECT DISTINCT setor FROM tb_ordens WHERE ordem_excluida IS NULL OR ordem_excluida = FALSE;")
        setores = cur.fetchall()

        cur.execute("SELECT * FROM tb_planejamento_anual")
        name_cols = ['codigo', 'tombamento', 'setor',
                     'descricao', 'criticidade', 'periodicidade']
        df_maquinas = pd.DataFrame(cur.fetchall()).iloc[:, :6]
        df_maquinas = df_maquinas.rename(
            columns=dict(zip(df_maquinas.columns, name_cols)))
        maquinas = df_maquinas.values.tolist()

        setor_selecionado = request.form.getlist('filtro_setor')

        lista_setore_selecionado = setor_selecionado

        if setor_selecionado:
            setor_selecionado = ', '.join([f"'{setor}'" for setor in setor_selecionado])
        else:
            setor_selecionado = None  # Ou qualquer valor padrão que faça sentido no seu caso

        # maquina_selecionado = request.form.get('filtro_maquinas')
        mes = request.form.get('datetimes')

        if not mes:
            mes_inicial_str = '2023-06-22'
            boleano_historico = False
            # Usar o construtor do datetime.date
            mes_inicial = datetime.date(datetime.strptime(mes_inicial_str, '%Y-%m-%d'))
            hoje = datetime.now().date()
            mes_final = hoje + timedelta(days=1)
        else:
            mes_inicial, mes_final = mes.split(' - ')
            mes_inicial = datetime.strptime(mes_inicial, '%d/%m/%Y').date()
            mes_final = datetime.strptime(mes_final, '%d/%m/%Y').date()

        # Criar uma lista de meses no intervalo
        lista_meses = []
        data_atual = mes_inicial

        while data_atual <= mes_final:
            lista_meses.append(data_atual.month)
            data_atual = data_atual.replace(day=1) + timedelta(days=32)
            data_atual = data_atual.replace(day=1)

        mes_inicial = mes_inicial.strftime('%Y-%m-%d')
        mes_final = mes_final.strftime('%Y-%m-%d')

        query = "SELECT * FROM tb_ordens WHERE 1=1"

        maquinas_importantes = request.form.getlist('maquinas-favoritas')

        if maquinas_importantes or len(maquinas) != 0:
            cur.execute(
                'SELECT DISTINCT (codigo) FROM tb_planejamento_anual')
            maquinas_preventivas = cur.fetchall()
            maquinas_preventivas = [valor[0] for valor in maquinas_preventivas]
            maquinas_selecionadas = ",".join(
                [f"'{maquinas}'" for maquinas in maquinas_preventivas])

        # Adiciona as condições de filtro se os campos não estiverem vazios
        if setor_selecionado:
            query += f" AND setor in ({setor_selecionado})"
        if mes_inicial:
            query += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query += f" AND maquina in ({maquinas_selecionadas})"

        query += ' AND ordem_excluida IS NULL OR ordem_excluida = FALSE;'

        # Executa a query
        cur.execute(query)
        itens_filtrados = cur.fetchall()

        # Criando cards

        # Monta a query base
        query = """
                SELECT *
                FROM tb_ordens
                WHERE (ordem_excluida IS NULL OR ordem_excluida = FALSE)
               """

        if mes_inicial:
            query += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if setor_selecionado:
            query += f" AND setor in ({setor_selecionado})"
        if maquinas_importantes:
            query += f" AND maquina in ({maquinas_selecionadas})"


        query_em_espera = """
                SELECT *
                FROM tb_ordens
                WHERE (ordem_excluida IS NULL OR ordem_excluida = FALSE)
               """

        if mes_inicial:
            query_em_espera += f" AND ultima_atualizacao >= '{mes_inicial}' AND ultima_atualizacao <= '{mes_final}'"
        if setor_selecionado:
            query_em_espera += f" AND setor in ({setor_selecionado})"
        if maquinas_importantes:
            query_em_espera += f" AND maquina in ({maquinas_selecionadas})"

        print("Query aqui", query)

        lista_qt = cards_post(query)

        card_em_espera = card_post_em_espera(query_em_espera)

        lista_qt.append(card_em_espera[0])

        print('lista_qt',lista_qt)

        """ Finalizando cards """

        """Criando gráficos de barras MTBF por maquina"""

        query_mtbf = (
            """
            SELECT 
                CASE 
                    WHEN t2.apelido IS NOT NULL AND t2.apelido <> '' THEN t2.apelido
                    ELSE t1.maquina
                END as maquina,
                t1.n_ordem,
                t1.id_ordem,
                t1.datafim,
                t1.setor
            FROM tb_ordens as t1
            LEFT JOIN tb_maquinas as t2 
                ON t1.maquina = t2.codigo
            WHERE 1=1 
        """)

        if setor_selecionado:
            query_mtbf += f" AND t1.setor in ({setor_selecionado})"
        if mes_inicial:
            query_mtbf += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query_mtbf += f" AND maquina in ({maquinas_selecionadas})"

        query_mtbf += " AND (t1.ordem_excluida IS NULL OR t1.ordem_excluida = FALSE) AND t1.natureza = 'OS';"

        """Criando gráficos de barras MTTR por maquina"""

        query_mttr = (
            """
            SELECT DISTINCT t3.id_ordem,status,datafim,maquina,t3.n_ordem,setor,inicio,fim,parada3
            FROM (
                SELECT
                    t1.id_ordem,
                    t1.status,
                    t1.datafim,
                    CASE
                        WHEN t2.apelido IS NOT NULL AND t2.apelido <> '' THEN t2.apelido
                        ELSE t1.maquina
                    END as maquina,
                    t1.n_ordem,
                    t1.setor,
                    TO_TIMESTAMP(t1.datainicio || ' ' || t1.horainicio, 'YYYY-MM-DD HH24:MI:SS') AS inicio,
                    TO_TIMESTAMP(t1.datafim || ' ' || t1.horafim, 'YYYY-MM-DD HH24:MI:SS') AS fim
                FROM tb_ordens as t1
                LEFT JOIN tb_maquinas as t2
                    ON t1.maquina = t2.codigo
                WHERE 1=1
        """)

        if setor_selecionado:
            query_mttr += f" AND t1.setor in ({setor_selecionado})"
        # if area_manutencao:
        #     query_mttr += f" AND area_manutencao = '{area_manutencao}'"
        if mes_inicial:
            query_mttr += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query_mttr += f" AND maquina in ({maquinas_selecionadas})"

        query_mttr += " AND (ordem_excluida IS NULL OR ordem_excluida = FALSE) AND natureza = 'OS') as t3 INNER JOIN tb_paradas t4 ON t3.id_ordem = t4.id_ordem and t3.n_ordem = t4.n_ordem;;"

        print(query_mttr)

        query_disponibilidade = ("""
            SELECT t1.*, t2.parada3
            FROM (
                SELECT
                    status,
                    id_ordem,
                    datafim,
                    maquina,
                    n_ordem,
                    setor,
                    TO_TIMESTAMP(datainicio || ' ' || horainicio, 'YYYY-MM-DD HH24:MI:SS') AS inicio,
                    TO_TIMESTAMP(datafim || ' ' || horafim, 'YYYY-MM-DD HH24:MI:SS') AS fim
                FROM
                    tb_ordens
                WHERE
                    1=1
            """)

        if setor_selecionado:
            query_disponibilidade += f" AND  setor in ({setor_selecionado})"
        # if area_manutencao:
        #     query_disponibilidade += f" AND area_manutencao = '{area_manutencao}'"
        if mes_inicial:
            query_disponibilidade += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query_disponibilidade += f" AND maquina in ({maquinas_selecionadas})"

        query_disponibilidade += """ AND (ordem_excluida IS NULL OR ordem_excluida = FALSE) ) AS t1 JOIN tb_paradas t2 ON t1.id_ordem = t2.id_ordem and t1.n_ordem = t2.n_ordem"""

        print(query_disponibilidade)

        query_horas_trabalhada_area = ("""
        SELECT
            area_manutencao,
            (CASE WHEN
                (SUM(horafim - horainicio) < INTERVAL '0') 
            THEN 
                (-SUM(horafim - horainicio)) 
            ELSE 
                SUM(horafim - horainicio)
            END) AS diferenca
        FROM tb_ordens
        WHERE 1=1
        """)
        
        if setor_selecionado:
            query_horas_trabalhada_area += f" AND setor in ({setor_selecionado})"
        if mes_inicial:
            query_horas_trabalhada_area += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query_horas_trabalhada_area += f" AND maquina in ({maquinas_selecionadas})"

        query_horas_trabalhada_area += " AND (ordem_excluida IS NULL OR ordem_excluida = FALSE) AND natureza = 'OS' GROUP BY area_manutencao;"

        print('query_horas_trabalhada_area',query_horas_trabalhada_area)

        query_horas_trabalhada_tipo = ("""
        SELECT
            tipo_manutencao,
            (CASE WHEN
                (SUM(horafim - horainicio) < INTERVAL '0') 
            THEN 
                (-SUM(horafim - horainicio)) 
            ELSE 
                SUM(horafim - horainicio)
            END) AS diferenca
        FROM tb_ordens
        WHERE 1=1
        """)

        if setor_selecionado:
            query_horas_trabalhada_tipo += f" AND setor in ({setor_selecionado})"
        if mes_inicial:
            query_horas_trabalhada_tipo += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query_horas_trabalhada_tipo += f" AND maquina in ({maquinas_selecionadas})"

        query_horas_trabalhada_tipo += " AND (ordem_excluida IS NULL OR ordem_excluida = FALSE) AND natureza = 'OS' GROUP BY tipo_manutencao;"

        query_horas_trabalhada_setor = ("""
        SELECT
            setor,
            (CASE WHEN
                (SUM(horafim - horainicio) < INTERVAL '0') 
            THEN 
                (-SUM(horafim - horainicio)) 
            ELSE 
                SUM(horafim - horainicio)
            END) AS diferenca
        FROM tb_ordens
        WHERE ordem_excluida IS NULL
        """)

        if setor_selecionado:
            query_horas_trabalhada_setor += f" AND setor in ({setor_selecionado})"
        if mes_inicial:
            query_horas_trabalhada_setor += f" AND datafim >= '{mes_inicial}' AND datafim <= '{mes_final}'"
        if maquinas_importantes:
            query_horas_trabalhada_setor += f" AND maquina in ({maquinas_selecionadas})"

        query_horas_trabalhada_setor += " AND (ordem_excluida IS NULL OR ordem_excluida = FALSE) AND natureza = 'OS' GROUP BY setor;"

        cur.execute(
            'SELECT DISTINCT EXTRACT(MONTH FROM ultima_atualizacao) AS numero_mes FROM tb_ordens;')

        """ Criando cards """

        resultado = funcao_geral(query_mtbf, query_mttr, boleano_historico, setor_selecionado,
                                query_disponibilidade, query_horas_trabalhada_tipo, query_horas_trabalhada_area, 
                                query_horas_trabalhada_setor, mes_inicial,mes_final,lista_meses)

        context_mtbf_maquina = resultado['context_mtbf_maquina']
        context_mttr_maquina = resultado['context_mttr_maquina']
        context_mttr_setor = resultado['context_mttr_setor']
        context_mtbf_setor = resultado['context_mtbf_setor']

        context_disponiblidade_maquina = resultado['context_disponibilidade']
        context_disponiblidade_setor = resultado['context_disponibilidade_setor']

        context_horas_trabalhadas = resultado['context_horas_trabalhadas']
        context_horas_trabalhadas_tipo = resultado['context_horas_trabalhadas_tipo']
        context_horas_trabalhadas_area = resultado['context_horas_trabalhadas_area']
        context_mtbf_top10_maquina = resultado['context_mtbf_top10_maquina']

        lista_horas_trabalhadas_area = resultado['lista_horas_trabalhadas_area']
        lista_horas_trabalhadas_tipo = resultado['lista_horas_trabalhadas_tipo']
        lista_horas_trabalhadas_setor = resultado['lista_horas_trabalhadas']

        lista_mtbf_setor = resultado['df_timeline_mtbf_setor']
        lista_mtbf_maquina = resultado['df_timeline_copia']
        
        lista_disponibilidade_setor = resultado['df_disponibilidade_setor']
        lista_disponibilidade_maquina = resultado['df_combinado_disponibilidade']
        lista_mttr_setor = resultado['df_combinado_mttr_setor']
        lista_mttr_maquina = resultado['df_combinado_mttr']
        top_10_maiores_MTBF_lista = resultado['top_10_maiores_MTBF_lista']
        disponibilidade_geral_maquina = resultado['disponibilidade_geral_maquina']
        disponibilidade_geral_setor = resultado['disponibilidade_geral_setor']

        return render_template('user/grafico.html', lista_qt=lista_qt,card_em_espera=card_em_espera, setores=setores, itens_filtrados=itens_filtrados,
                               lista_setore_selecionado=lista_setore_selecionado, **context_mtbf_maquina,
                               **context_mtbf_setor, **context_mttr_maquina, **context_mttr_setor, **context_disponiblidade_maquina, **context_horas_trabalhadas_area, **context_horas_trabalhadas_tipo,
                               **context_mtbf_top10_maquina, **context_disponiblidade_setor, mes=mes, **context_horas_trabalhadas, lista_horas_trabalhadas_setor=lista_horas_trabalhadas_setor,
                               lista_horas_trabalhadas_tipo=lista_horas_trabalhadas_tipo, lista_horas_trabalhadas_area=lista_horas_trabalhadas_area, lista_mtbf_setor=lista_mtbf_setor,
                               lista_mtbf_maquina=lista_mtbf_maquina, lista_disponibilidade_setor=lista_disponibilidade_setor,
                               top_10_maiores_MTBF_lista=top_10_maiores_MTBF_lista, lista_disponibilidade_maquina=lista_disponibilidade_maquina,
                               lista_mttr_setor=lista_mttr_setor, lista_mttr_maquina=lista_mttr_maquina,
                               todos_meses=todos_meses, maquinas_importantes=maquinas_importantes, disponibilidade_geral_maquina=disponibilidade_geral_maquina,
                               disponibilidade_geral_setor=disponibilidade_geral_setor)

    mes = datetime.now().month
    mes = list(range(1, mes + 1))

    boleano_historico = True
    setor_selecionado = None

    # Monta a query base
    query = """
            SELECT *
            FROM tb_ordens
            WHERE (ordem_excluida IS NULL OR ordem_excluida = FALSE)
            """

    lista_qt = cards_get(query)

    query_mtbf = (
        """
        SELECT maquina, n_ordem, id_ordem, datafim, setor
        FROM tb_ordens
        WHERE 1=1 AND ordem_excluida IS NULL OR ordem_excluida = FALSE AND natureza = 'OS'
    """)

    query_mttr = (
        """
        SELECT t1.id_ordem,t1.n_ordem,t1.setor,t1.maquina,
            TO_TIMESTAMP(datainicio || ' ' || horainicio, 'YYYY-MM-DD HH24:MI:SS') AS inicio,
            TO_TIMESTAMP(datafim || ' ' || horafim, 'YYYY-MM-DD HH24:MI:SS') AS fim,
            parada3
        FROM tb_ordens as t1
        left join tb_paradas as t2 on t1.id_ordem = t2.id_ordem and t1.n_ordem = t2.n_ordem
        WHERE 1=1 AND ordem_excluida IS NULL OR ordem_excluida = FALSE AND natureza = 'OS'
    """)

    query_disponibilidade = (
        """
        SELECT t1.*, t2.parada3
        FROM (
            SELECT
                datafim, 
                maquina, 
                n_ordem,
                setor,
                status,
                id_ordem,
                TO_TIMESTAMP(datainicio || ' ' || horainicio, 'YYYY-MM-DD HH24:MI:SS') AS inicio,
                TO_TIMESTAMP(datafim || ' ' || horafim, 'YYYY-MM-DD HH24:MI:SS') AS fim
            FROM
                tb_ordens
            WHERE
                1=1
            AND (ordem_excluida IS NULL OR ordem_excluida = FALSE)) AS t1 JOIN tb_paradas t2 ON t1.id_ordem = t2.id_ordem and t1.n_ordem = t2.n_ordem
    """)

    query_horas_trabalhada_tipo = """
        SELECT
            tipo_manutencao,
            (CASE WHEN
                (SUM(horafim - horainicio) < INTERVAL '0') 
            THEN 
                (-SUM(horafim - horainicio)) 
            ELSE 
                SUM(horafim - horainicio)
            END) AS diferenca
        FROM tb_ordens
        WHERE ordem_excluida IS NULL
        GROUP BY tipo_manutencao;
        """

    query_horas_trabalhada_area = """
        SELECT
            area_manutencao,
            (CASE WHEN
                (SUM(horafim - horainicio) < INTERVAL '0') 
            THEN 
                (-SUM(horafim - horainicio)) 
            ELSE 
                SUM(horafim - horainicio)
            END) AS diferenca
        FROM tb_ordens
        WHERE ordem_excluida IS NULL
        GROUP BY area_manutencao;
        """
    
    query_horas_trabalhada_setor = """
        SELECT
            setor,
            (CASE WHEN
                (SUM(horafim - horainicio) < INTERVAL '0') 
            THEN 
                (-SUM(horafim - horainicio)) 
            ELSE 
                SUM(horafim - horainicio)
            END) AS diferenca
        FROM tb_ordens
        WHERE ordem_excluida IS NULL
        GROUP BY setor;
        """
    
    data_inicial = datetime(2022,8,1).date() # Data Inicial do Histórico do MTBF
    data_fim = datetime.now().date() # Data Final do Histórico do MTBF,sempre hoje

    resultado = funcao_geral(query_mtbf, query_mttr, boleano_historico, setor_selecionado,
                            query_disponibilidade, query_horas_trabalhada_tipo, query_horas_trabalhada_area,
                            query_horas_trabalhada_setor,data_inicial,data_fim, mes)

    context_mtbf_maquina = resultado['context_mtbf_maquina']
    context_mttr_maquina = resultado['context_mttr_maquina']
    context_mttr_setor = resultado['context_mttr_setor']
    context_mtbf_setor = resultado['context_mtbf_setor']

    context_disponiblidade_maquina = resultado['context_disponibilidade']
    context_disponiblidade_setor = resultado['context_disponibilidade_setor']

    context_horas_trabalhadas = resultado['context_horas_trabalhadas']
    context_horas_trabalhadas_tipo = resultado['context_horas_trabalhadas_tipo']
    context_horas_trabalhadas_area = resultado['context_horas_trabalhadas_area']
    context_mtbf_top10_maquina = resultado['context_mtbf_top10_maquina']

    lista_horas_trabalhadas_area = resultado['lista_horas_trabalhadas_area']
    lista_horas_trabalhadas_tipo = resultado['lista_horas_trabalhadas_tipo']
    lista_horas_trabalhadas_setor = resultado['lista_horas_trabalhadas']

    lista_mtbf_setor = resultado['df_timeline_mtbf_setor']
    lista_mtbf_maquina = resultado['df_timeline_copia']
    lista_disponibilidade_setor = resultado['df_disponibilidade_setor']
    lista_disponibilidade_maquina = resultado['df_combinado_disponibilidade']
    lista_mttr_setor = resultado['df_combinado_mttr_setor']
    lista_mttr_maquina = resultado['df_combinado_mttr']
    top_10_maiores_MTBF_lista = resultado['top_10_maiores_MTBF_lista']
    disponibilidade_geral_maquina = resultado['disponibilidade_geral_maquina']
    disponibilidade_geral_setor = resultado['disponibilidade_geral_setor']

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute(
        "SELECT DISTINCT setor FROM tb_ordens WHERE ordem_excluida IS NULL OR ordem_excluida = FALSE;")
    setores = cur.fetchall()

    cur.execute("SELECT * FROM tb_planejamento_anual")
    name_cols = ['codigo', 'tombamento', 'setor',
                 'descricao', 'criticidade', 'periodicidade']
    df_maquinas = pd.DataFrame(cur.fetchall()).iloc[:, :6]
    df_maquinas = df_maquinas.rename(
        columns=dict(zip(df_maquinas.columns, name_cols)))
    maquinas = df_maquinas.values.tolist()

    # Se o método for GET, exibe todos os itens
    cur.execute(
        "SELECT * FROM tb_ordens WHERE ordem_excluida IS NULL OR ordem_excluida = FALSE")
    itens = cur.fetchall()

    cur.close()
    conn.close()

    mes_descrito = obter_nome_mes(mes)

    return render_template('user/grafico.html', lista_qt=lista_qt, setores=setores, maquinas=maquinas, itens=itens, mes_descrito=mes_descrito,
                           **context_mtbf_maquina, **context_mtbf_setor, **context_mttr_maquina, **context_mttr_setor, **context_mtbf_top10_maquina,
                           **context_disponiblidade_maquina, **context_disponiblidade_setor, **context_horas_trabalhadas, **context_horas_trabalhadas_tipo,
                           **context_horas_trabalhadas_area, lista_horas_trabalhadas_area=lista_horas_trabalhadas_area, lista_horas_trabalhadas_tipo=lista_horas_trabalhadas_tipo,
                           top_10_maiores_MTBF_lista=top_10_maiores_MTBF_lista, lista_mtbf_setor=lista_mtbf_setor, lista_mtbf_maquina=lista_mtbf_maquina, lista_horas_trabalhadas_setor=lista_horas_trabalhadas_setor,
                           setor_selecionado='', lista_disponibilidade_setor=lista_disponibilidade_setor, lista_disponibilidade_maquina=lista_disponibilidade_maquina,
                           lista_mttr_setor=lista_mttr_setor, lista_mttr_maquina=lista_mttr_maquina, area_manutencao='', disponibilidade_geral_maquina=disponibilidade_geral_maquina,
                           disponibilidade_geral_setor=disponibilidade_geral_setor)



@routes_bp.route('/timeline', methods=['POST'])
@login_required
def timeline_os():

    dados = request.get_json()
    print(dados)

    id_ordem = dados['id_ordem']
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = ("""
        SELECT
            o.dataabertura,
            o.n_ordem,
            o.status,
            o.operador,
            o.descmanutencao,
			COALESCE(TO_TIMESTAMP(o.datainicio || ' ' || o.horainicio, 'YYYY-MM-DD HH24:MI:SS'), TO_TIMESTAMP(o.dataabertura  - INTERVAL '3 hours' || ' ' || '00:00:00', 'YYYY-MM-DD HH24:MI:SS')) AS inicio,
			COALESCE(TO_TIMESTAMP(o.datafim || ' ' || o.horafim, 'YYYY-MM-DD HH24:MI:SS'), TO_TIMESTAMP(o.dataabertura  - INTERVAL '3 hours' || ' ' || '00:00:00', 'YYYY-MM-DD HH24:MI:SS')) AS fim,
            func.nome,
            func.matricula,
            func.salario
        FROM tb_ordens as o
		LEFT JOIN tb_funcionario as func ON ',' || o.operador || ',' LIKE '%,' || func.matricula || ',%'
        WHERE o.id_ordem = {} AND (o.ordem_excluida IS NULL OR o.ordem_excluida = FALSE);
        """).format(int(id_ordem))
    
    print(s)

    df_timeline = pd.read_sql_query(s, conn)
    print(df_timeline)
    
    # df_timeline['inicio'] = df_timeline['inicio'].apply(lambda x: x.split(' ')[0])
    # df_timeline['fim'] = df_timeline['fim'].apply(lambda x: x.split(' ')[0])


    # for i in range(len(df_timeline)):
    #     if df_timeline['fim'][i] == 'NaT':
    #         df_timeline['fim'][i] = 0
    #         df_timeline['inicio'][i] = 0
    #     else:
    #         pass

    df_timeline = df_timeline.replace(np.nan, '-')
    
    try:
        df_timeline['inicio'] = pd.to_datetime(df_timeline['inicio'])
        df_timeline['fim'] = pd.to_datetime(df_timeline['fim'])

        # df_timeline['diferenca'] = pd.to_datetime(df_timeline['fim']) - pd.to_datetime(df_timeline['inicio'])
        df_timeline['diferenca'] = (df_timeline['fim'] - df_timeline['inicio']).apply(
            lambda x: x.total_seconds() // 60 if pd.notnull(x) else None)

        for i in range(len(df_timeline)):
            df_timeline['operador'][i] = df_timeline['operador'][i].replace(
                "{", "").replace("[", "").replace("\\", "").replace('"', '').replace("]}", "")
    except:
        df_timeline['diferenca'] = 0

    df_timeline = df_timeline.sort_values(by='n_ordem', ascending=True)

    if df_timeline['inicio'][0] == '-':
        df_timeline['inicio'][0] = df_timeline['dataabertura'][0]

    df_final = df_timeline

    print(df_timeline)

    if len(df_timeline) > 1:
        df_timeline['mesExecucao'] = df_timeline['fim'].dt.month
        df_timeline['anoExecucao'] = df_timeline['fim'].dt.year
        df_timeline['dias_uteis'] = df_timeline.apply(
            lambda row: calcular_dias_uteis(int(row['anoExecucao']), int(row['mesExecucao'])), axis=1)
        df_timeline['horasTotalMes'] = df_timeline['dias_uteis'] * (9*60)
        df_timeline['salario'] = df_timeline['salario'].replace("-", 0)
        df_timeline['salario'] = df_timeline['salario'].astype(float)
        df_timeline['proporcional'] = (
            df_timeline['salario'] * df_timeline['diferenca']) / df_timeline['horasTotalMes']

        df_groupby = df_timeline[['n_ordem', 'proporcional']].groupby(
            ['n_ordem']).sum().reset_index().round(2)

        df_timeline = df_timeline.drop(columns=[
                                       'mesExecucao', 'anoExecucao', 'dias_uteis', 'horasTotalMes', 'proporcional', 'nome', 'matricula', 'salario'])
        df_timeline = df_timeline.drop_duplicates(subset=['n_ordem'])

        df_final = pd.merge(df_timeline, df_groupby, how='left', on='n_ordem')

        df_final['diferenca'] = df_final['diferenca'].astype(int)

        totalMinutos = df_final['diferenca'].sum().tolist()
        totalCusto = df_final['proporcional'].sum().round(2).tolist()

        print(totalMinutos, totalCusto)

        df_final = df_final.iloc[:, 1:]

        df_final = df_final.values.tolist()

        return jsonify (id_ordem, df_final, totalMinutos, totalCusto)
    else:
        df_final = df_final.iloc[:, 1:]

        df_final = df_final.values.tolist()

        return jsonify (id_ordem, df_final)

def tabela_maquinas_preventivas():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = (""" SELECT * FROM tb_planejamento_anual """)

    cur.execute(s)
    tabela_maquinas_preventivas = cur.fetchall()

    return tabela_maquinas_preventivas

@routes_bp.route('/tabela-grupos-preventivas')
@login_required
def grupos_preventivas():
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    s = (""" select * from tb_grupos_preventivas where excluidos = 'False' """)

    data = pd.read_sql_query(s, conn)

    # Retirando os grupos que não tem data estabelecida
    df_grupos_nan = data[data.isna().any(axis=1)]
    df_grupos_nan['proxima_manutencao'] = 'À decidir'
    df_grupos_nan.fillna('À decidir',inplace=True)
    
    # Grupos que ja tem data estabelecida
    df_grupos_notna = data.dropna()
    df_grupos_notna['proxima_manutencao'] = df_grupos_notna.apply(lambda row: calcular_proxima_data(row['ult_manutencao'], float(row['periodicidade'])*30), axis=1)

    df_grupos_notna['proxima_manutencao'] = pd.to_datetime(df_grupos_notna['proxima_manutencao'],format="%Y-%m-%d").dt.strftime("%d/%m/%Y")
    df_grupos_notna['ult_manutencao'] = pd.to_datetime(df_grupos_notna['ult_manutencao'],format="%Y-%m-%d").dt.strftime("%d/%m/%Y")

    # Juntando os dois grupos
    grupos_juntos = pd.concat([df_grupos_notna,df_grupos_nan]).values.tolist()

    return jsonify(grupos_juntos)

@routes_bp.route('/tabela-historico-preventivas')
@login_required
def historico_planejadas():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query_historico_preventivas = """select max(n_ordem) as n_ordem_max,datafim,id_ordem,maquina, coalesce(status,'Em espera') as status 
                                    from tb_ordens 
                                    where natureza = 'Planejada' and ordem_excluida isnull
                                    group by status,datafim,id_ordem,maquina
                                    order by id_ordem,n_ordem_max"""

    data_historico_planejadas = pd.read_sql_query(query_historico_preventivas,conn)

    data_historico_planejadas.drop_duplicates(subset='id_ordem',keep='last',inplace=True)
    # data_historico_planejadas['datafim'] = data_historico_planejadas['datafim'].fillna('-')

    data_historico_planejadas['datafim'] = pd.to_datetime(data_historico_planejadas['datafim'],format="%Y-%m-%d",errors='coerce').dt.strftime("%d/%m/%Y").fillna('-')

    data_historico_planejadas = data_historico_planejadas.values.tolist()

    return jsonify(data_historico_planejadas)


@routes_bp.route('/52semanas', methods=['GET'])
@login_required
def plan_52semanas():

    # tabela com máquinas que podem ser criadas grupos de preventiva
    tabela_maquinas_preventivas_ = tabela_maquinas_preventivas()

    return render_template('user/52semanas.html', tabela_maquinas_preventivas_=tabela_maquinas_preventivas_)

@routes_bp.route('/preventivas', methods=['GET'])
@login_required
def preventivas():
    
    """
    Rota para visualizar as opções de atividades preventivas
    """

    # Obter o código da máquina a partir dos parâmetros da consulta
    codigo_maquina = request.args.get('codigo_maquina')

    # Use o código da máquina para gerar as opções dinamicamente
    # Substitua esta lógica pela lógica real que você precisa
    opcoes = obter_opcoes_preventivas(codigo_maquina)

    return jsonify(opcoes)

# Função de exemplo para gerar opções com base no código da máquina
def obter_opcoes_preventivas(codigo_maquina):

    """
    Função para buscar grupos de atividades por máquina caso tenha
    """
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = f"""SELECT DISTINCT (grupo) FROM tb_grupos_preventivas WHERE codigo = '{codigo_maquina}' AND excluidos = 'false' """

    cur.execute(sql)
    grupos = cur.fetchall()

    if len(grupos) == 0:
        return [[]]
    else:
        return grupos

@routes_bp.route('/atividadesGrupo', methods=['GET'])
@login_required
def atividadesGrupo():
    # Obtenha os parâmetros da consulta
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    codigo_maquina = request.args.get('codigo_maquina')
    grupo_selecionado = request.args.get('grupo')

    sql = 'SELECT ult_manutencao,periodicidade FROM tb_grupos_preventivas WHERE codigo = %s and grupo = %s'

    cur.execute(sql,(codigo_maquina,grupo_selecionado))
    data = cur.fetchall()

    # Use os parâmetros para carregar os dados associados
    dados_associados, parametros = tarefasGrupo(codigo_maquina, grupo_selecionado)  
    
    print(data)

    try:
        nova_data = data[0][0].strftime("%Y-%m-%d")
        periodicidade = data[0][1]
    except:
        nova_data = None
        periodicidade = None
    
    df = pd.DataFrame({'data': [nova_data],
                    'periodicidade': [periodicidade]})
    df.index = [0]  # Adiciona um índice à primeira linha

    try:
        df['data'] = pd.to_datetime(df['data'])
    except:
        pass

    try:
        df['proxima_manutencao'] = df.apply(lambda row: calcular_proxima_data(row['data'], float(row['periodicidade'])*30), axis=1)
        proxima_data = df['proxima_manutencao'][0]
    except Exception as e:
        print(f"Erro ao calcular próxima manutenção: {e}")
        proxima_data = None

    print(parametros)

    if len(parametros)>0:
        parametros[0].append(formatar_data(proxima_data))
    else:
        parametros = None


    # Retorne os dados como JSON
    return jsonify(dados_associados,parametros)

# Função de exemplo para obter dados associados a uma máquina e grupo

def tarefasGrupo(codigo_maquina, grupo_selecionado):
    
    """
    Função para buscar atividades preventivas de acordo com o grupo escolhido
    """
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    sql = f"""SELECT * FROM tb_atividades_preventiva WHERE codigo = '{codigo_maquina}' and grupo = '{grupo_selecionado}' and excluidos = 'false' """

    cur.execute(sql)
    atividades = cur.fetchall()
        
    sql = f"""SELECT ult_manutencao,periodicidade FROM tb_grupos_preventivas WHERE codigo = '{codigo_maquina}' and grupo = '{grupo_selecionado}'"""

    cur.execute(sql)
    parametros = cur.fetchall()

    if len(parametros) > 0:
        parametros[0][0] = formatar_data(parametros[0][0])

    if len(atividades) == 0:
        return [[]],parametros
    else:
        return atividades,parametros

@routes_bp.route("/excluir-tarefa", methods=['POST'])
def excluirTarefa():
    
    """
    Rota para excluir uma atividade preventiva
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()
    codigo_maquina = data['codigo_maquina']
    grupo = data['grupoSelecionado']
    excluidos = True
    try:
        idDaLinha = data['idDaLinha']
    except:
        idDaLinha = None

    if idDaLinha == '' or idDaLinha == None:
        sql_delete = f"""UPDATE public.tb_atividades_preventiva SET excluidos = '{excluidos}' WHERE codigo = '{codigo_maquina}' AND grupo = '{grupo}'; """

    else:
        sql_delete = f"""DELETE FROM public.tb_atividades_preventiva WHERE grupo = '{grupo}' and id = '{idDaLinha}' """

    cur.execute(sql_delete)

    conn.commit()
    conn.close()

    return 'sucess'


def proxima_data_util(data_inicial, periodicidade):
  
  """
  Retorna a próxima data útil, contando apenas dias úteis, sem sábados e domingos.

  Args:
    data_inicial: A data inicial.
    periodicidade: A periodicidade da data.

  Returns:
    A próxima data útil.
  """

  # Converte as datas para objetos do tipo date.

  data_inicial = date.fromisoformat(data_inicial)

  # Calcula o número de dias úteis entre a data inicial e a data atual.

  dias_uteis = 0
  while data_inicial <= date.today():
    if data_inicial.weekday() < 5:
      dias_uteis += 1
    data_inicial += timedelta(days=1)

  # Adiciona a periodicidade à data inicial.

  data_final = data_inicial + timedelta(days=int(periodicidade))

  # Verifica se a data final é um sábado ou domingo.

  if data_final.weekday() >= 5:
    # Empurra a data para segunda-feira.
    data_final += timedelta(days=3)

  return data_final.isoformat()


def formatar_data(data):
    
    """
    Função para formatar data dentro da lista
    """

    return data.strftime("%Y-%m-%d") if isinstance(data, date) else data


@routes_bp.route('/criar-grupo', methods=['POST'])
def rota_criar_grupo():

    """
    Rota para receber o nome da máquina e criar o grupo
    """

    data = request.get_json()

    nome_grupo = data['nome_grupo']
    codigo_maquina = data['codigo_maquina']

    resultado_criar_grupo = criar_grupo(codigo_maquina,nome_grupo)

    if resultado_criar_grupo == "Grupo já existente":
        return jsonify("Grupo já existente")

    return 'sucess'

def criar_grupo(codigo_maquina,nome_grupo):

    """
    Função para criar grupo de atividades preventivas ao clicar no botão "Criar grupo".
    Recebe apenas a máquina e busca o último grupo criado a ela e adicionar + 1.
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = f"""SELECT * FROM tb_grupos_preventivas WHERE codigo = '{codigo_maquina}' AND grupo = '{nome_grupo}' """

    cur.execute(sql)
    grupos = cur.fetchall()

    if len(grupos) > 0:
        return "Grupo já existente"
        
    sql = f"""INSERT INTO tb_grupos_preventivas (codigo,grupo) VALUES ('{codigo_maquina}','{nome_grupo}')"""

    cur.execute(sql)

    conn.commit()
    conn.close()

    return "Sucesso"


@routes_bp.route('/receber-tarefas', methods=['POST'])
def receber_tarefas():

    """
    Rota para receber as atividades associadas a maquina e ao grupo
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    json_tarefas = request.get_json()
    print(json_tarefas)

    periodicidade = float(json_tarefas['parametros'][0]['periodicidade_grupo'])
    ultima_manutencao = json_tarefas['parametros'][0]['ultima_manutencao']
    grupo = json_tarefas['parametros'][0]['grupo']
    codigo_maquina = json_tarefas['parametros'][0]['codigo_maquina']

    sql_update = f"""UPDATE tb_grupos_preventivas SET ult_manutencao = '{ultima_manutencao}', periodicidade = {periodicidade}
        WHERE codigo = '{codigo_maquina}' and grupo = '{grupo}'"""

    sql_insert = f"""INSERT INTO tb_grupos_preventivas (codigo,grupo,ult_manutencao,periodicidade) VALUES ('{codigo_maquina}', '{grupo}',
        '{ultima_manutencao}', '{periodicidade}')"""
    
    try:
        cur.execute(sql_update)
    except:
        cur.execute(sql_insert) 

    conn.commit()

    if len(json_tarefas['dadosTabela']) > 0:
        adicionar_editar_tarefa(json_tarefas)
  
    return 'sucess'


def adicionar_editar_tarefa(json_tarefas):
    """
    Função para editar e adicionar tarefa no banco de dados.
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    for tarefa in range(len(json_tarefas['dadosTabela'])):
        
        if json_tarefas['dadosTabela'][tarefa]['atividadeAntiga'] != '':

            atividade_atual = json_tarefas['dadosTabela'][tarefa]['atividade']
            responsabilidade_atual = json_tarefas['dadosTabela'][tarefa]['responsabilidade']

            atividade_antiga = json_tarefas['dadosTabela'][tarefa]['atividadeAntiga']
            responsabilidade_antiga = json_tarefas['dadosTabela'][tarefa]['responsabilidadeAntiga']

            id_unico = int(json_tarefas['dadosTabela'][tarefa]['id_unico'])

            sql_edit = f"""UPDATE tb_atividades_preventiva SET atividade = '{atividade_atual}', responsabilidade = '{responsabilidade_atual}', id = {id_unico}
                        WHERE atividade = '{atividade_antiga}' and responsabilidade = '{responsabilidade_antiga}' and id = {id_unico} """
            
            cur.execute(sql_edit)
        
        else:

            atividade = json_tarefas['dadosTabela'][tarefa]['atividade'] 
            responsabilidade = json_tarefas['dadosTabela'][tarefa]['responsabilidade']
            grupo = json_tarefas['dadosTabela'][tarefa]['grupo']
            codigo_maquina = json_tarefas['dadosTabela'][tarefa]['codigo_maquina']
            
            sql = f"""
                    INSERT INTO tb_atividades_preventiva (codigo,grupo,responsabilidade,atividade)
                        VALUES ('{codigo_maquina}','{grupo}','{responsabilidade}','{atividade}')
                    """

            cur.execute(sql)

    conn.commit()
    conn.close()


@routes_bp.route('/receber-upload', methods=['POST'])
def receber_upload():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Obter o arquivo do formulário
    file = request.files['file']

    # Obter outras informações do formulário
    grupo_selecionado = request.form['grupoSelecionado']
    codigo_maquina = request.form['codigo_maquina']   

    print(file)

    print(codigo_maquina)

    # Salvar o arquivo no servidor (opcional)
    # file.save('uploads_atividade/' + file.filename)

    # file = r"uploads_atividade/" + file.filename

    try:
        df = pd.read_csv(file, sep=";", encoding='ISO-8859-1')
    except pd.errors.ParserError:
        df = pd.read_excel(file)

    # Remover caracteres especiais do cabeçalho das colunas
    df.columns = df.columns.str.replace('ï»¿', '')

    colunas_esperadas = ['codigo_maquina', 'responsabilidade', 'atividade']  # Substitua com as colunas reais
    

    # Verificar se as colunas do DataFrame coincidem com as colunas esperadas
    if set(df.columns) != set(colunas_esperadas):
        return 'Colunas do arquivo não correspondem ao modelo'

    df['grupo'] = grupo_selecionado

    df = df[['codigo_maquina','grupo','responsabilidade','atividade']]

    df_list = df.values.tolist()

    print(df_list)

    for row in df_list:       

        codigo_maquina_grupo = row[0]
        if codigo_maquina != codigo_maquina_grupo:
            return 'Código de máquina inválido'
        grupo = row[1]
        responsabilidade = row[2]
        atividade = row[3]

        sql_insert = f"""INSERT INTO tb_atividades_preventiva (codigo,grupo,responsabilidade,atividade)
                        VALUES ('{codigo_maquina_grupo}','{grupo}','{responsabilidade}','{atividade}')"""

        cur.execute(sql_insert)

    conn.commit()
    conn.close()

    # os.remove(file)
    
    return 'sucess'


@routes_bp.route('/download-modelo-atividades', methods=['GET'])
def download_modelo_excel():
    # Caminho para o arquivo modelo CSV
    excel_filename = 'modelo_atividades.csv'

    # Envie o arquivo para download
    return send_file(excel_filename, as_attachment=True)


@routes_bp.route('/cadastrar52', methods=['POST', 'GET'])
@login_required
def cadastro_preventiva():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == 'POST':

        try:
            togglePreventiva = request.form.get('cadastrar-preventiva')
            codigo = request.form['codigo']
            tombamento = request.form['tombamento']
            descricao = request.form['descricao']
            setor = request.form['setor']
            criticidade = request.form['criticidade']
            manut_inicial = request.form['manut_inicial']
            periodicidade = request.form['periodicidade']
            apelido = request.form['apelido']

            print(codigo,setor,descricao,tombamento,criticidade,manut_inicial,periodicidade,togglePreventiva)

            # df = gerador_de_semanas_informar_manutencao(
            #     setor, codigo, descricao, tombamento, criticidade, manut_inicial, periodicidade)

            print("antes")

            # codigo = 'teste1'
            # setor = 'Administrativo'
            # descricao ='123'
            # tombamento = '123'
            # criticidade = 'B'
            # manut_inicial = '2023-12-12'
            # periodicidade = 30
            # apelido = 'teste'

            df = gerar_planejamento_maquinas_preventivas(codigo,setor,descricao,tombamento,criticidade,manut_inicial,periodicidade)
            
            print("depois")

            df['ultima_manutencao'] = df['ultima_manutencao'].dt.strftime("%Y-%m-%d")
            df['proxima_manutencao'] = df['proxima_manutencao'].dt.strftime("%Y-%m-%d")
            df['periodicidade'] = df['periodicidade'].astype(str)

            lista = df.values.tolist()
            # lista = lista[0]

            print(lista)

            s = ("""
                SELECT * FROM tb_planejamento_anual
                """)

            maquina_cadastrada = pd.read_sql_query(s, conn)

            if len(maquina_cadastrada[maquina_cadastrada['codigo'] == codigo]) > 0:
                flash("Máquina ja cadastrada", category='danger')

            else:

                try:
                    conn = psycopg2.connect(
                        dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
                    cur = conn.cursor(
                        cursor_factory=psycopg2.extras.DictCursor)

                    # Consulta SQL para inserir os dados na tabela
                    sql_insert = "INSERT INTO tb_planejamento_anual VALUES ({})".format(','.join(['%s'] * len(lista[0])))

                    # Executar a consulta SQL para cada sublista
                    for linha in lista:
                        cur.execute(sql_insert, linha)

                    query_max = ("""SELECT max(id) FROM tb_maquinas""")
                    cur.execute(query_max)
                    id = cur.fetchall()
                    id = id[0][0] + 1

                    cur.execute("INSERT INTO tb_maquinas (id, setor, codigo, descricao, tombamento,apelido) VALUES (%s,%s, %s, %s, %s,%s)",
                                (id, setor, codigo, descricao, tombamento, apelido))

                    # Confirmar a transação
                    conn.commit()

                    flash("Máquina cadastrada com sucesso", category='sucess')

                except Error as e:
                    print(
                        f"Ocorreu um erro ao conectar ou executar a consulta no PostgreSQL: {e}")

                finally:
                    # Fechar o cursor e a conexão com o banco de dados
                    cur.close()
                    conn.close()

                flash("Máquina cadastrada com sucesso", category='sucess')

            return render_template('user/cadastrar52.html')

        except:
            codigo = request.form['codigo']
            tombamento = request.form['tombamento']
            descricao = request.form['descricao']
            setor = request.form['setor']
            apelido = request.form['apelido']

            s = ("""
                SELECT * FROM tb_maquinas
                """)

            query_max = ("""SELECT max(id) FROM tb_maquinas""")
            cur.execute(query_max)
            id = cur.fetchall()
            id = id[0][0] + 1

            maquina_cadastrada = pd.read_sql_query(s, conn)

            if len(maquina_cadastrada[maquina_cadastrada['codigo'] == codigo]) > 0:
                flash("Máquina ja cadastrada", category='danger')

            else:

                cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

                cur.execute("INSERT INTO tb_maquinas (id, setor, codigo, descricao, tombamento,apelido) VALUES (%s,%s, %s, %s, %s,%s)",
                            (id, setor, codigo, descricao, tombamento, apelido))

                conn.commit()

                flash("Máquina cadastrada com sucesso", category='sucess')

            return render_template('user/cadastrar52.html')

    return render_template('user/cadastrar52.html')

@routes_bp.route('/verificar-codigo-existente', methods=['POST'])
def verificar_codigo_existente():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    codigo = request.get_json()

    print(codigo)

    cur.execute('select codigo from tb_maquinas where codigo = %s', (codigo,))

    data = cur.fetchall()

    if len(data)>0:
        return jsonify({'codigo_existente':True})
    else:
        return jsonify({'codigo_existente':False})

def adicionar_maquina(dados):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute('insert into public.tb_maquinas (setor,codigo,descricao,tombamento,apelido) values(%s,%s,%s,%s,%s)',(dados['setor'],dados['codigo'],dados['descricao'],
                                                                                                                     dados['tombamento'],dados['apelido']))

    conn.commit()

    return 'sucess'

def adicionar_maquina_preventiva(dados):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    cur.execute('insert into public.tb_planejamento_anual (codigo,classificacao) values(%s,%s)',(dados['codigo'],dados['criticidade']))

    conn.commit()

    return 'sucess'

@routes_bp.route('/cadastrar-maquina', methods=['POST'])
@login_required
def cadastrar_maquina():

    dados = request.get_json()
    
    adicionar_maquina(dados)

    if dados['preventiva']:
        adicionar_maquina_preventiva(dados)

    return 'sucess'


@routes_bp.route('/testes_envio_pdf/<codigo_maquina>', methods=['POST'])
@login_required
def testes_envio_pdf(codigo_maquina):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor()

    # Certifique-se de usar 'pdfFile' para corresponder ao nome do campo no FormData
    pdfs = request.files.getlist('pdfFile')

    print(pdfs)

    if len(pdfs) > 0:
        for pdf in pdfs:
            if pdf.filename != '':
                pdf_data = pdf.read()
                pdf_filename = pdf.filename  # Obtém o nome do arquivo

                cur.execute("INSERT INTO tb_anexos (codigo_maquina, checklist, nome_arquivo) VALUES (%s,%s,%s)",
                            (codigo_maquina, pdf_data, pdf_filename))
                conn.commit()

    cur.close()
    conn.close()

    # Você pode personalizar a mensagem de retorno conforme necessário
    return jsonify({"message": "Upload bem-sucedido"})


@routes_bp.route('/visualizar_midias/<id_ordem>', methods=['GET'])
@login_required
def visualizar_midias(id_ordem):
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor()

    # Buscar as imagens associadas à ordem de serviço
    cur.execute("SELECT imagem FROM tb_imagens WHERE id_ordem = %s", (id_ordem,))
    imagens_data = [base64.b64encode(row[0]).decode(
        'utf-8') for row in cur.fetchall()]

    return jsonify(imagens_data=imagens_data, id_ordem=id_ordem)


@routes_bp.route('/visualizar_video/<id_ordem>', methods=['GET'])
@login_required
def visualizar_video(id_ordem):
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor()

    # Buscar os vídeos associados à ordem de serviço
    cur.execute(
        "SELECT video FROM tb_videos_ordem_servico WHERE id_ordem = %s", (id_ordem,))
    videos_data = [base64.b64encode(row[0]).decode('utf-8')
                   for row in cur.fetchall()]

    # Convertendo os dados de vídeo em URLs
    video_urls = []
    for video_data in videos_data:
        video_url = f"data:video/mp4;base64,{video_data}"
        video_urls.append(video_url)

    return jsonify(videos_data=video_urls)


@routes_bp.route('/timeline-preventiva/<maquina>', methods=['POST', 'GET'])
@login_required
# Mostrar o histórico de preventiva daquela máquina
def timeline_preventiva(maquina):

    print(maquina)

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Obtém os dados da tabela
    s = ("""
        SELECT * 
        FROM tb_ordens
        WHERE ordem_excluida IS NULL OR ordem_excluida = FALSE
        """)

    df = pd.read_sql_query(s, conn)
    df['maquina'] = df['maquina'].str.strip()

    df = df[df['maquina'] == maquina].reset_index(drop=True)
    # df = df[df['natureza'] == 'Planejada'].reset_index(drop=True)

    df[['dataabertura', 'id_ordem']]

    # Limpar a coluna
    for i in range(len(df)):
        try:
            df['operador'][i] = df['operador'][i].replace("{", "").replace("}", "").replace(
                "[", "").replace("]", "").replace("\\", "").replace('"', '').replace("]}", "").replace("}}", "")
        except:
            pass

    # Criar um dicionário para mapear cada ID à sua respectiva data
    id_data_map = {}

    # Iterar sobre os IDs únicos e encontrar a data correspondente para cada um
    for id_ordem in df['id_ordem'].unique():
        data = df.loc[df['id_ordem'] == id_ordem, 'dataabertura'].iloc[0]
        id_data_map[id_ordem] = data

    # Atualizar os valores de dataabertura para cada ID
    df['dataabertura'] = df['id_ordem'].map(id_data_map)

    df = df.drop_duplicates(subset='id_ordem', keep='last')

    df = df.sort_values('id_ordem', ascending=True)

    data = df.values.tolist()

    return render_template('user/timeline_preventiva.html', data=data, maquina=maquina)


@routes_bp.route('/mostrar_pdf/<codigo_maquina>', methods=['GET'])
@login_required
def mostrar_pdf(codigo_maquina):

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor()

    cur.execute(
        "SELECT checklist, nome_arquivo FROM tb_anexos WHERE codigo_maquina = %s", (codigo_maquina,))
    pdf_records = cur.fetchall()

    pdf_urls = []
    nome_arquivos = []

    for pdf_record in pdf_records:
        pdf_data, nome_arquivo = pdf_record
        pdf_stream = BytesIO(pdf_data)

        # Gere o URL do download incluindo o nome do arquivo
        pdf_url = f'/download_pdf/{codigo_maquina}/{len(pdf_urls) + 1}'
        pdf_urls.append(pdf_url)
        nome_arquivos.append(nome_arquivo)

    cur.close()
    conn.close()

    return jsonify({'pdfUrls': pdf_urls, 'nome_arquivo': nome_arquivos})


@routes_bp.route('/download_pdf/<codigo_maquina>/<int:pdf_index>', methods=['GET'])
@login_required
def download_pdf(codigo_maquina, pdf_index):
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor()

    print(pdf_index)

    cur.execute(
        "SELECT checklist FROM tb_anexos WHERE codigo_maquina = %s", (codigo_maquina,))
    pdf_records = cur.fetchall()

    if 0 <= pdf_index - 1 < len(pdf_records):
        pdf_data = pdf_records[pdf_index - 1][0]
        pdf_stream = BytesIO(pdf_data)

        cur.close()
        conn.close()

        response = make_response(pdf_stream.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=pdf_{pdf_index}.pdf'
        return response
    else:
        cur.close()
        conn.close()
        return jsonify({'error': 'PDF não encontrado'}), 404


@routes_bp.route('/remover_pdf/<codigo_maquina>/<nome_arquivo>', methods=['DELETE'])
@login_required
def remover_pdf(codigo_maquina, nome_arquivo):
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor()

    print(nome_arquivo)

    cur.execute("SELECT id FROM tb_anexos WHERE codigo_maquina = %s AND nome_arquivo = %s",
                (codigo_maquina, nome_arquivo))
    pdf_id = cur.fetchone()

    if pdf_id:
        cur.execute("DELETE FROM tb_anexos WHERE id = %s", (pdf_id[0],))
        conn.commit()

        cur.close()
        conn.close()

        return jsonify({"message": "PDF removido com sucesso"})
    else:
        cur.close()
        conn.close()
        return jsonify({'error': 'PDF não encontrado'}), 404


@routes_bp.route('/lista_maquinas', methods=['GET'])
@login_required
def lista_maquinas():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cur.execute(""" SELECT 
                        tb_planejamento_anual.codigo,
                        tb_maquinas.setor,
                        tb_maquinas.descricao,
                        tb_maquinas.tombamento,
                        tb_maquinas.apelido
                    FROM tb_planejamento_anual
                    JOIN tb_maquinas ON tb_planejamento_anual.codigo = tb_maquinas.codigo; """)

    df_c_preventivas = pd.DataFrame(cur.fetchall(), columns=[
                                    'codigo', 'setor', 'descricao', 'tombamento', 'apelido'])
    df_c_preventivas['setor'] = df_c_preventivas['setor'].str.title()
    df_c_preventivas['preventiva'] = 'Y'

    cur.execute(""" SELECT codigo, setor, descricao, tombamento, apelido
                    FROM tb_maquinas; """)

    df_s_preventivas = pd.DataFrame(cur.fetchall(), columns=[
                                    'codigo', 'setor', 'descricao', 'tombamento', 'apelido'])
    df_s_preventivas['setor'] = df_s_preventivas['setor'].str.title()
    df_s_preventivas['preventiva'] = 'N'

    df_final = pd.concat([df_c_preventivas, df_s_preventivas]).drop_duplicates(
        subset='codigo', keep='first').reset_index(drop=True)

    for i in range(len(df_final)):
        if df_final['tombamento'][i] == None:
            df_final['tombamento'][i] = ''
        if df_final['apelido'][i] == None:
            df_final['apelido'][i] = ''

    data = df_final.values.tolist()

    return render_template('user/lista_maquinas.html', data=data)


@routes_bp.route('/excluir-ordem', methods=['POST'])
@login_required
def excluir_ordem():
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()
    id_linha = data['id']
    texto = data['texto']

    print(id_linha, texto)

    query = """UPDATE tb_ordens
            SET ordem_excluida = 'yes', motivo_exclusao = %s
            WHERE id_ordem = %s
            """

    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cur.execute(query, [texto, id_linha])
    conn.commit()

    return 'Dados recebidos com sucesso!'


@routes_bp.route('/visualizar_pdf/<id_ordem>')
@login_required
def visualizar_pdf(id_ordem):

    return formulario_os(id_ordem)


@routes_bp.route('/transformar-maquina',methods=['POST'])
@login_required
def transformar_maquina():

    data = request.get_json()

    codigo_maquina = data['codigo_preventivas']
    classificacao = data['classificacao']

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    print(codigo_maquina,classificacao)
    
    cur.execute("INSERT INTO tb_planejamento_anual (codigo, classificacao) VALUES (%s, %s)",
                            (codigo_maquina, classificacao))
    print("Máquina transformada com sucesso")

    conn.commit()
    conn.close()

    return 'Sucesso'


@routes_bp.route('/modal-editar-maquina',methods=['POST'])
@login_required
def editar_maquina():

    codigo_maquina = request.get_json()

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    query = """SELECT 
                    tb_maquinas.codigo,
                    tb_maquinas.setor,
                    tb_maquinas.apelido,
                    tb_maquinas.descricao,
                    tb_maquinas.tombamento
                FROM tb_maquinas
                WHERE tb_maquinas.codigo = '{}';""".format(codigo_maquina)

    cur.execute(query)
    data = cur.fetchall()

    setor = data[0][1]
    apelido = data[0][2]
    descricao = data[0][3]
    tombamento = data[0][4]

    if not tombamento:
        tombamento = ''

    if not apelido:
        apelido = ''

    data = {'codigo':codigo_maquina,'setor':setor,'descricao':descricao,'tombamento':tombamento,'apelido':apelido}

    return jsonify(data)


@routes_bp.route('/envio-edicao-maquina-preventiva', methods=['POST'])
@login_required
def envio_editar_maquina_preventiva():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()

    codigo_maquina = data['codigo_preventivas']
    tombamento = data['tombamento_preventivas']
    descricao = data['descricao_preventivas']
    apelido = data['apelido_preventivas']

    print(tombamento,descricao,apelido)

    cur.execute("""
                UPDATE tb_maquinas
                SET tombamento=%s,descricao=%s,apelido=%s
                WHERE codigo = %s
                """, (tombamento, descricao, apelido, codigo_maquina))

    conn.commit()
    conn.close()
    
    return jsonify(codigo_maquina)

@routes_bp.route('/modal-editar-maquina-preventiva', methods=['POST'])
@login_required
def editar_maquina_preventiva():

    codigo_maquina = request.get_json()

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """SELECT 
                tb_planejamento_anual.codigo,
                tb_maquinas.tombamento,
                tb_maquinas.setor,
                tb_maquinas.descricao,
                tb_planejamento_anual.classificacao,
                tb_maquinas.apelido
            FROM tb_planejamento_anual
            JOIN tb_maquinas ON tb_planejamento_anual.codigo = tb_maquinas.codigo
            WHERE tb_planejamento_anual.codigo = '{}';""".format(codigo_maquina)

    cur.execute(query)
    data = cur.fetchall()

    codigo_maquina = codigo_maquina
    tombamento = data[0][1]
    setor = data[0][2]
    descricao = data[0][3]
    criticidade = data[0][4]
    apelido = data[0][5]

    if not tombamento:
        tombamento = ''
    if not apelido:
        apelido = ''

    data = {'codigo':codigo_maquina,'setor':setor,'descricao':descricao,'tombamento':tombamento,'apelido':apelido,
            'criticidade':criticidade}

    return jsonify(data)


@routes_bp.route('/envio-edicao-maquina', methods=['POST'])
@login_required
def envio_editar_maquina():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()

    codigo_maquina = data['codigo']
    tombamento = data['tombamento']
    descricao = data['descricao']
    apelido = data['apelido']

    print(tombamento,descricao,apelido,codigo_maquina)

    cur.execute("""
                UPDATE tb_maquinas
                SET tombamento=%s,descricao=%s,apelido=%s
                WHERE codigo = %s
                """, (tombamento, descricao, apelido, codigo_maquina))

    conn.commit()
    conn.close()
    
    return 'Sucesso'


@routes_bp.route('/excluir-maquina', methods=['POST'])
@login_required
def excluir_maquina():
    if request.method == 'POST':
        # Obter o código da máquina enviado pelo frontend
        codigo_maquina = request.form.get('codigo_maquina')

        conn = psycopg2.connect(
            dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        query = """DELETE FROM tb_maquinas
                WHERE codigo = %s;
                """

        cur.execute(query, [codigo_maquina])

        query = """DELETE FROM tb_planejamento_anual
                WHERE codigo = %s;
                """

        cur.execute(query, [codigo_maquina])

        conn.commit()
        conn.close()

        # flash("Máquina excluída com sucesso", category='sucess')

        return 'Dados recebidos com sucesso!'


@routes_bp.route('/excluir-preventiva', methods=['POST'])
@login_required
def excluir_preventiva():

    if request.method == 'POST':
        # Obter o código da máquina enviado pelo frontend
        codigo_maquina = request.form.get('codigo_maquina')

        conn = psycopg2.connect(
            dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        query = """DELETE FROM tb_planejamento_anual
                WHERE codigo = %s;
                """

        cur.execute(query, [codigo_maquina])

        conn.commit()
        conn.close()

        # flash("Máquina excluída com sucesso", category='sucess')

        return 'Dados recebidos com sucesso!'


@routes_bp.route('/excluir-execucao', methods=['POST'])
@login_required
def excluir_execucao():

    id_ordem = int(request.form.get('id_ordem'))
    n_ordem = int(request.form.get('n_execucao'))

    print(id_ordem, n_ordem)

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """
            DELETE FROM tb_ordens
            WHERE id_ordem=%s and n_ordem=%s;
            """

    cur.execute(query, [id_ordem, n_ordem])

    conn.commit()
    conn.close()

    flash("Execução excluída com sucesso", category='sucess')

    return 'Execução excluída com sucesso'

@routes_bp.route('/excluir-grupo', methods=['POST'])
@login_required
def excluir_grupo():


    json_grupos_excluidos = request.get_json()

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    excluido = True
    
    codigo_maquina = json_grupos_excluidos['codigo_maquina']

    grupoSelecionado = json_grupos_excluidos['grupoSelecionado']

    print(codigo_maquina,grupoSelecionado,excluido)

    cur.execute(""" UPDATE tb_grupos_preventivas
                    SET excluidos=%s
                    WHERE codigo = %s AND grupo = %s
                    """, (excluido, codigo_maquina, grupoSelecionado))

    conn.commit()
    conn.close()

    flash("Execução excluída com sucesso", category='sucess')

    return 'Execução excluída com sucesso'

@routes_bp.route("/funcionarios", methods=['POST', 'GET'])
@login_required
def funcionarios():
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == 'POST':
        nome = request.form['nome']
        matricula = request.form['matricula']
        ativo = request.form['ativo']
        salario = request.form['salario']
        funcao = request.form['funcao']

        print(nome, matricula, ativo, salario, funcao)

        s = (""" SELECT * FROM tb_funcionario""")

        funcionario_cadastrado = pd.read_sql_query(s, conn)

        if len(funcionario_cadastrado[funcionario_cadastrado['nome'] == nome]) > 0 or len(funcionario_cadastrado[funcionario_cadastrado['matricula'] == matricula]) > 0:
            flash("Funcionário ja cadastrado", category='danger')
            print("Funcionário ja cadastrado")
        else:

            cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

            cur.execute("INSERT INTO tb_funcionario (nome, matricula, ativo, salario, funcao) VALUES (%s, %s, %s, %s,%s)",
                        (nome, matricula, ativo, salario, funcao))
            print("Funcionário cadastrado com sucesso")

            conn.commit()
            conn.close()

            flash("Funcionário cadastrado com sucesso", category='sucess')
            return render_template('user/funcionarios.html')

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """SELECT * FROM tb_funcionario"""

    cur.execute(query)
    data = cur.fetchall()
    df_data = pd.DataFrame(data)
    funci = df_data.values.tolist()

    return render_template("user/funcionarios.html", funci=funci)


@routes_bp.route("/editar_funcionarios", methods=['POST', 'GET'])
@login_required
def editar_funcionarios():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    if request.method == 'POST':
        nome_antigo = request.form.get('nome_antigo')
        nome_novo = request.form.get('nome')
        matricula = request.form.get('matricula')
        ativo = request.form.get('ativo')
        salario = request.form.get('salario')
        funcao = request.form.get('funcao')

        # Use o valor selecionado no <select> para atualizar o registro no banco de dados
        query = """
        UPDATE tb_funcionario
        SET nome = %s, matricula = %s, ativo = %s, salario = %s, funcao = %s
        WHERE nome = %s
        """

        cur.execute(query, (nome_novo, matricula, ativo,
                    salario, funcao, nome_antigo))
        conn.commit()

        return render_template("user/funcionarios.html")

    selected_value = request.args.get('selectedValue')
    if selected_value:
        query = """SELECT nome, matricula, ativo, salario, funcao FROM tb_funcionario WHERE nome = '{}';""".format(
            selected_value)
        cur.execute(query)
        data = cur.fetchone()  # Assume que há apenas um registro correspondente
    else:
        # Se nenhum valor foi selecionado, retorne valores vazios
        data = {'nome': '', 'matricula': '',
                'ativo': '', 'salario': '', 'funcao': ''}

    return jsonify(data)